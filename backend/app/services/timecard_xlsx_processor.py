"""
Serviço para processar arquivos XLSX de Cartão Ponto
"""
import logging
import unicodedata
import openpyxl
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Tuple
from sqlalchemy import or_
from sqlalchemy.orm import Session
from sqlalchemy.exc import ProgrammingError, SQLAlchemyError

from app.core.database import Base, engine
from app.models.employee import Employee
from app.models.timecard import TimecardPeriod, TimecardData, TimecardProcessingLog

logger = logging.getLogger(__name__)


class TimecardXLSXProcessor:
    """Processador de arquivos XLSX de cartão ponto"""

    HEADER_SCAN_LIMIT = 25

    REQUIRED_COLUMNS = ['employee_number', 'employee_name']

    COLUMN_ALIASES = {
        'employee_number': ['Nº Folha', 'N° Folha', 'N Folha', 'Nro Folha', 'Matrícula', 'Matricula', 'Folha'],
        'employee_name': ['Nome', 'Colaborador', 'Funcionário', 'Funcionario'],
        'normal_hours': ['Normais', 'Horas Normais'],
        'overtime_50': ['Ex50%', 'HE50%', 'HE 50%', 'Hora Extra 50%'],
        'overtime_100': ['Ex100%', 'HE100%', 'HE 100%', 'Hora Extra 100%'],
        'night_overtime_50': ['EN50%', 'EN 50%'],
        'night_overtime_100': ['EN100%', 'EN 100%'],
        # IMPORTANTE: night_hours é APENAS para "Adicional Noturno" / "Noturno" (não DSR)
        'night_hours': ['Adic. Noturno', 'Adicional Noturno', 'Noturno', 'Not.', 'Adic Noturno', 'Adic. Not.'],
        'absences': ['Faltas', 'Intrajornada', 'Intra Jornada', 'Horas Intrajornada', 'Interj.'],
        # IMPORTANTE: dsr_debit - priorizar "DSR" (coluna que tem os valores reais ~29h/pessoa)
        # depois "DSR.Deb" (específico de débito, frequentemente vazio)
        'dsr_debit': ['DSR', 'DSR.Deb', 'DSR Deb', 'DSR Débito', 'DSR Déb', 'DSR.Débito'],
        'bonus_hours': ['Abono2', 'Abono 2']
    }
    
    def __init__(self, db: Session, user_id: Optional[int] = None):
        self.db = db
        self.user_id = user_id
        self.warnings = []
        self.errors = []
    
    def process_xlsx_file(
        self,
        file_path: str,
        year: int,
        month: int,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        dry_run: bool = False,
    ) -> Dict:
        """
        Processa arquivo XLSX de cartão ponto
        
        Args:
            file_path: Caminho do arquivo XLSX
            year: Ano do período
            month: Mês do período
            start_date: Data início (opcional)
            end_date: Data fim (opcional)
        
        Returns:
            Dict com resultado do processamento
        """
        start_time = datetime.now()
        
        try:
            logger.info(f"Iniciando processamento de cartão ponto: {file_path}")
            
            # Carregar workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            
            # Encontrar linha de headers
            header_row = self._find_header_row(sheet)
            if not header_row:
                return {
                    'success': False,
                    'error': 'Não foi possível encontrar os headers das colunas'
                }
            
            # Ler e validar headers
            headers = self._read_headers(sheet, header_row)
            column_indices = self._resolve_column_indices(headers)
            validation_result = self._validate_headers(column_indices)
            if not validation_result['valid']:
                return {
                    'success': False,
                    'error': validation_result['error']
                }
            
            period = None
            if dry_run:
                period = self._get_existing_period(year, month)
            else:
                # Criar ou obter período
                period = self._get_or_create_period(
                    year, month, start_date, end_date, file_path
                )
            
            # Processar linhas de dados
            results = self._process_data_rows(
                sheet,
                column_indices,
                header_row,
                period,
                file_path,
                year,
                month,
                dry_run=dry_run,
            )
            
            processing_time = (datetime.now() - start_time).total_seconds()

            if not dry_run and period is not None:
                # Criar log de processamento
                self._create_processing_log(
                    period, file_path, results, processing_time
                )
                
                # Commit final
                self.db.commit()
            
            wb.close()
            
            logger.info(f"Processamento concluído: {results['processed_rows']} linhas processadas")

            if dry_run:
                return {
                    'success': True,
                    'dry_run': True,
                    'period_name': self._build_period_name(year, month),
                    'would_create_period': results.get('would_create_period', False),
                    'period_exists': results.get('period_exists', False),
                    'total_rows': results['total_rows'],
                    'processed_rows': results['processed_rows'],
                    'error_rows': results['error_rows'],
                    'matched_by_matricula': results.get('matched_by_matricula', 0),
                    'matched_by_name': results.get('matched_by_name', 0),
                    'unmatched_rows': results.get('unmatched_rows', 0),
                    'would_create_records': results.get('would_create_records', 0),
                    'would_update_records': results.get('would_update_records', 0),
                    'headers_used': results.get('headers_used', []),
                    'warnings': self.warnings,
                    'errors': self.errors,
                    'processing_time': processing_time,
                }
            
            return {
                'success': True,
                'period_id': period.id,
                'period_name': period.period_name,
                'total_rows': results['total_rows'],
                'processed_rows': results['processed_rows'],
                'error_rows': results['error_rows'],
                'warnings': self.warnings,
                'errors': self.errors,
                'processing_time': processing_time
            }
            
        except Exception as e:
            self.db.rollback()
            logger.error(f"Erro ao processar XLSX de cartão ponto: {e}", exc_info=True)
            return {
                'success': False,
                'error': f'Erro ao processar arquivo: {str(e)}'
            }
    
    def _find_header_row(self, sheet) -> Optional[int]:
        """Localiza a linha de cabeçalho usando aliases conhecidos."""
        best_row = None
        best_score = 0

        for row_idx in range(1, min(sheet.max_row, self.HEADER_SCAN_LIMIT) + 1):
            headers = self._read_headers(sheet, row_idx)
            column_indices = self._resolve_column_indices(headers)
            score = len(column_indices)

            if score >= 2 and all(required in column_indices for required in self.REQUIRED_COLUMNS):
                logger.info(f"Headers encontrados na linha {row_idx}")
                return row_idx

            if score > best_score:
                best_row = row_idx
                best_score = score

        if best_row is not None and best_score >= 2:
            logger.info(f"Usando linha {best_row} como header por maior correspondência")
            return best_row

        return None
    
    def _read_headers(self, sheet, header_row: int) -> List[str]:
        """Lê os headers da linha especificada"""
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(header_row, col).value
            headers.append(str(cell_value).strip() if cell_value else None)
        return headers
    
    def _normalize_header_value(self, value: Optional[str]) -> str:
        if value is None:
            return ''

        normalized = unicodedata.normalize('NFKD', str(value))
        normalized = normalized.encode('ascii', 'ignore').decode('ascii')
        normalized = normalized.lower()
        normalized = ''.join(ch if ch.isalnum() else ' ' for ch in normalized)
        normalized = ' '.join(normalized.split())
        return normalized

    def _resolve_column_indices(self, headers: List[str]) -> Dict[str, int]:
        """Resolve os cabeçalhos reais do XLSX para nomes canônicos internos.

        A resolução é conservadora:
        1. match exato com o texto original do header
        2. match case-insensitive
        3. fallback por normalização apenas para acento/espaço

        Isso preserva o nome real da coluna do XLSX e evita que campos sem
        header correspondente usem posições fixas por engano.
        """
        column_indices: Dict[str, int] = {}
        header_map = {header: index + 1 for index, header in enumerate(headers) if header}

        normalized_headers: Dict[str, int] = {}
        lowered_headers: Dict[str, int] = {}
        for header, index in header_map.items():
            lowered_headers[header.lower()] = index
            normalized_headers[self._normalize_header_value(header)] = index

        for canonical_name, aliases in self.COLUMN_ALIASES.items():
            candidate_headers = [canonical_name, *aliases]

            for candidate in candidate_headers:
                if candidate in header_map:
                    column_indices[canonical_name] = header_map[candidate]
                    break

                lowered_candidate = candidate.lower()
                if lowered_candidate in lowered_headers:
                    column_indices[canonical_name] = lowered_headers[lowered_candidate]
                    break

                normalized_candidate = self._normalize_header_value(candidate)
                if normalized_candidate in normalized_headers:
                    column_indices[canonical_name] = normalized_headers[normalized_candidate]
                    break

        return column_indices

    def _validate_headers(self, column_indices: Dict[str, int]) -> Dict:
        """Valida se as colunas necessárias estão presentes."""
        for required_col in self.REQUIRED_COLUMNS:
            if required_col not in column_indices:
                display_name = self.COLUMN_ALIASES.get(required_col, [required_col])[0]
                return {
                    'valid': False,
                    'error': f'Coluna obrigatória não encontrada: {display_name}'
                }
        return {'valid': True}

    def _build_period_name(self, year: int, month: int) -> str:
        month_names = [
            'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
        ]
        return f"{month_names[month - 1]} {year}"

    def _get_existing_period(self, year: int, month: int) -> Optional[TimecardPeriod]:
        try:
            return self.db.query(TimecardPeriod).filter(
                TimecardPeriod.year == year,
                TimecardPeriod.month == month,
            ).first()
        except ProgrammingError as exc:
            if 'timecard_periods' in str(exc):
                self.db.rollback()
                return None
            raise

    def _ensure_timecard_tables(self) -> None:
        """Cria as tabelas de cartão ponto se ainda não existirem."""
        Base.metadata.create_all(
            bind=engine,
            tables=[
                TimecardPeriod.__table__,
                TimecardData.__table__,
                TimecardProcessingLog.__table__,
            ],
            checkfirst=True,
        )

    def _normalize_person_name(self, value: Optional[str]) -> str:
        """Normaliza nomes para comparação segura de identidade."""
        if not value:
            return ''
        return self._normalize_header_value(value)

    def _is_name_consistent(self, imported_name: Optional[str], employee_name: Optional[str]) -> bool:
        """Valida se nome importado e nome do cadastro pertencem à mesma pessoa.

        Regra principal: igualdade após normalização.
        Fallback leve: contém (para casos com sufixos/apelidos no cadastro).
        """
        imported = self._normalize_person_name(imported_name)
        stored = self._normalize_person_name(employee_name)

        if not imported or not stored:
            return False

        if imported == stored:
            return True

        # Fallback conservador para pequenas variações semânticas.
        return imported in stored or stored in imported

    def _resolve_employee(self, employee_number: str, employee_name: str, company: str):
        """Resolve colaborador com validação forte de matrícula + nome.

        Prioridade:
        1) Candidatos por matrícula/unique_id e validação de nome consistente
        2) Fallback por nome somente quando não houver candidato por matrícula
        """
        employee_number_clean = ''.join(ch for ch in employee_number if ch.isdigit())
        employee_number_nozeros = employee_number_clean.lstrip('0') or employee_number_clean

        registration_candidates = {
            employee_number_clean,
            employee_number_nozeros,
            employee_number_clean.zfill(5),
        }

        unique_candidates = {
            employee_number_clean,
            employee_number_nozeros,
            f"{company}{employee_number_clean}",
            f"{company}{employee_number_nozeros}",
            f"{company}{employee_number_clean.zfill(5)}",
        }

        employees = self.db.query(Employee).filter(
            or_(
                Employee.registration_number.in_(list(registration_candidates)),
                Employee.unique_id.in_(list(unique_candidates)),
            )
        ).all()

        if employees:
            # Se temos nome no arquivo, exige consistência nome + matrícula.
            if employee_name:
                name_consistent = [
                    emp for emp in employees
                    if self._is_name_consistent(employee_name, emp.name)
                ]

                if len(name_consistent) == 1:
                    return name_consistent[0], 'matricula_nome'

                # Ambíguo ou inconsistente: não vincula automaticamente.
                return None, None

            # Sem nome no arquivo: só vincula automaticamente se houver candidato único.
            if len(employees) == 1:
                return employees[0], 'matricula'

            return None, None

        if employee_name:
            employee = self.db.query(Employee).filter(
                Employee.name.ilike(f"%{employee_name}%")
            ).first()
            if employee:
                return employee, 'nome'

        return None, None

    def _get_cell_value(self, sheet, row_idx: int, col_indices: Dict[str, int], field_name: str):
        column_index = col_indices.get(field_name)
        if not column_index:
            return None
        return sheet.cell(row_idx, column_index).value

    def _extract_timecard_row_data(self, sheet, row_idx: int, col_indices: Dict[str, int]) -> Optional[Dict]:
        employee_number_raw = self._get_cell_value(sheet, row_idx, col_indices, 'employee_number')
        employee_number = str(employee_number_raw).strip().upper() if employee_number_raw is not None else ''
        employee_number_clean = ''.join(ch for ch in employee_number if ch.isdigit())

        if not employee_number_clean:
            return None

        company = '0060' if employee_number.endswith('E') else '0059'
        employee_name_cell = self._get_cell_value(sheet, row_idx, col_indices, 'employee_name')
        employee_name = str(employee_name_cell).strip() if employee_name_cell else ''

        normal_hours_cell = self._get_cell_value(sheet, row_idx, col_indices, 'normal_hours')
        overtime_50_cell = self._get_cell_value(sheet, row_idx, col_indices, 'overtime_50')
        overtime_100_cell = self._get_cell_value(sheet, row_idx, col_indices, 'overtime_100')
        night_overtime_50_cell = self._get_cell_value(sheet, row_idx, col_indices, 'night_overtime_50')
        night_overtime_100_cell = self._get_cell_value(sheet, row_idx, col_indices, 'night_overtime_100')
        night_hours_cell = self._get_cell_value(sheet, row_idx, col_indices, 'night_hours')
        absences_cell = self._get_cell_value(sheet, row_idx, col_indices, 'absences')
        dsr_debit_cell = self._get_cell_value(sheet, row_idx, col_indices, 'dsr_debit')
        bonus_hours_cell = self._get_cell_value(sheet, row_idx, col_indices, 'bonus_hours')

        return {
            'employee_number': employee_number,
            'employee_number_clean': employee_number_clean,
            'company': company,
            'employee_name': employee_name,
            'normal_hours': self._convert_to_hours(normal_hours_cell),
            'overtime_50': self._convert_to_hours(overtime_50_cell),
            'overtime_100': self._convert_to_hours(overtime_100_cell),
            'night_overtime_50': self._convert_to_hours(night_overtime_50_cell),
            'night_overtime_100': self._convert_to_hours(night_overtime_100_cell),
            'night_hours': self._convert_to_hours(night_hours_cell),
            'absences': self._convert_to_hours(absences_cell),
            'dsr_debit': self._convert_to_hours(dsr_debit_cell),
            'bonus_hours': self._convert_to_hours(bonus_hours_cell),
        }

    def _preview_timecard_row(self, row_data: Dict, period: Optional[TimecardPeriod]) -> Dict:
        employee, match_type = self._resolve_employee(
            employee_number=row_data['employee_number'],
            employee_name=row_data['employee_name'],
            company=row_data['company'],
        )

        existing = None
        if period is not None:
            existing = self.db.query(TimecardData).filter(
                TimecardData.period_id == period.id,
                TimecardData.employee_number == row_data['employee_number']
            ).first()

        return {
            'employee_found': bool(employee),
            'employee_id': employee.id if employee else None,
            'match_type': match_type,
            'would_create': existing is None,
            'would_update': existing is not None,
        }
    
    def _get_or_create_period(
        self,
        year: int,
        month: int,
        start_date: Optional[str],
        end_date: Optional[str],
        filename: str
    ) -> TimecardPeriod:
        """Cria ou retorna período existente"""
        try:
            self._ensure_timecard_tables()

            # Buscar período existente
            period = self.db.query(TimecardPeriod).filter(
                TimecardPeriod.year == year,
                TimecardPeriod.month == month
            ).first()
            
            if period:
                logger.info(f"Período existente encontrado: {period.period_name}")
                return period
            
            # Criar novo período
            month_names = [
                'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
            ]
            period_name = f"{month_names[month-1]} {year}"
            
            # Converter datas se fornecidas
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
            
            period = TimecardPeriod(
                year=year,
                month=month,
                period_name=period_name,
                start_date=start_date_obj,
                end_date=end_date_obj,
                description=f"Importado de {filename}"
            )
            
            self.db.add(period)
            self.db.flush()  # Get ID without committing
            
            logger.info(f"Novo período criado: {period_name}")
            return period
            
        except Exception as e:
            logger.error(f"Erro ao criar período: {e}")
            raise
    
    def _process_data_rows(
        self,
        sheet,
        column_indices: Dict[str, int],
        header_row: int,
        period: Optional[TimecardPeriod],
        filename: str,
        year: int,
        month: int,
        dry_run: bool = False,
    ) -> Dict:
        """Processa todas as linhas de dados"""
        total_rows = 0
        processed_rows = 0
        error_rows = 0
        matched_by_matricula = 0
        matched_by_name = 0
        unmatched_rows = 0
        would_create_records = 0
        would_update_records = 0
        
        # Processar linhas (arquivo tratado não tem linha de totais)
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            # Verificar se é linha de dados válida
            row_data = self._extract_timecard_row_data(sheet, row_idx, column_indices)

            if not row_data:
                continue  # Linha vazia
            
            total_rows += 1
            
            try:
                if dry_run:
                    preview = self._preview_timecard_row(row_data, period)
                    if preview['match_type'] in ('matricula', 'matricula_nome'):
                        matched_by_matricula += 1
                    elif preview['match_type'] == 'nome':
                        matched_by_name += 1

                    if not preview['employee_found']:
                        unmatched_rows += 1

                    if preview['would_create']:
                        would_create_records += 1
                    if preview['would_update']:
                        would_update_records += 1

                    processed_rows += 1
                else:
                    success, match_type = self._process_timecard_row(
                        row_data=row_data,
                        period=period,
                        filename=filename,
                    )

                    if success:
                        processed_rows += 1
                        if match_type in ('matricula', 'matricula_nome'):
                            matched_by_matricula += 1
                        elif match_type == 'nome':
                            matched_by_name += 1
                        elif match_type is None:
                            unmatched_rows += 1
                    else:
                        error_rows += 1
                    
            except Exception as e:
                error_rows += 1
                error_msg = f"Linha {row_idx}: {str(e)}"
                self.errors.append(error_msg)
                logger.error(error_msg)
        
        return {
            'total_rows': total_rows,
            'processed_rows': processed_rows,
            'error_rows': error_rows,
            'matched_by_matricula': matched_by_matricula,
            'matched_by_name': matched_by_name,
            'unmatched_rows': unmatched_rows,
            'would_create_records': would_create_records,
            'would_update_records': would_update_records,
            'period_exists': period is not None,
            'would_create_period': dry_run and period is None,
            'headers_used': list(column_indices.keys()),
        }
    
    def _process_timecard_row(
        self,
        row_data: Dict,
        period: TimecardPeriod,
        filename: str
    ) -> Tuple[bool, Optional[str]]:
        """Processa uma linha de dados de cartão ponto"""
        try:
            employee, match_type = self._resolve_employee(
                employee_number=row_data['employee_number'],
                employee_name=row_data['employee_name'],
                company=row_data['company'],
            )

            employee_id = employee.id if employee else None

            if not employee:
                warning = f"Colaborador '{row_data['employee_name']}' (matrícula {row_data['employee_number']}) não encontrado no sistema"
                self.warnings.append(warning)
                logger.warning(warning)
            
            # Verificar se já existe registro para este colaborador neste período
            existing = self.db.query(TimecardData).filter(
                TimecardData.period_id == period.id,
                TimecardData.employee_number == row_data['employee_number']
            ).first()
            
            if existing:
                # Atualizar registro existente
                existing.employee_id = employee_id
                existing.employee_name = row_data['employee_name']
                existing.company = row_data['company']
                existing.normal_hours = row_data['normal_hours']
                existing.overtime_50 = row_data['overtime_50']
                existing.overtime_100 = row_data['overtime_100']
                existing.night_overtime_50 = row_data['night_overtime_50']
                existing.night_overtime_100 = row_data['night_overtime_100']
                existing.night_hours = row_data['night_hours']
                existing.absences = row_data['absences']
                existing.dsr_debit = row_data['dsr_debit']
                existing.bonus_hours = row_data['bonus_hours']
                existing.upload_filename = filename
                existing.processed_by = self.user_id
                
                logger.debug(f"Registro atualizado: {row_data['employee_name']}")
            else:
                # Criar novo registro
                timecard_data = TimecardData(
                    period_id=period.id,
                    employee_id=employee_id,
                    employee_number=row_data['employee_number'],
                    employee_name=row_data['employee_name'],
                    company=row_data['company'],
                    normal_hours=row_data['normal_hours'],
                    overtime_50=row_data['overtime_50'],
                    overtime_100=row_data['overtime_100'],
                    night_overtime_50=row_data['night_overtime_50'],
                    night_overtime_100=row_data['night_overtime_100'],
                    night_hours=row_data['night_hours'],
                    absences=row_data['absences'],
                    dsr_debit=row_data['dsr_debit'],
                    bonus_hours=row_data['bonus_hours'],
                    upload_filename=filename,
                    processed_by=self.user_id
                )
                
                self.db.add(timecard_data)
                logger.debug(f"Novo registro criado: {row_data['employee_name']}")
            
            return True, match_type
            
        except Exception as e:
            logger.error(f"Erro ao processar linha: {e}")
            raise
    
    def _convert_to_hours(self, value) -> Decimal:
        """
        Converte valor para horas decimais
        
        Suporta:
        - timedelta objects (ex: timedelta(days=3, seconds=58260))
        - Números decimais do Excel (dias como fração)
        - Strings de tempo (ex: "3 days, 16:11:00" ou "1:21:00" ou "3153:13")
        """
        if value is None or value == "":
            return Decimal('0')
        
        # Se for timedelta do Python
        if isinstance(value, timedelta):
            total_seconds = value.total_seconds()
            # Converter para horas com precisão de 4 casas decimais para preservar minutos
            # Exemplo: 2737:26 = 2737.4333 horas (não arredondar para 2 casas)
            hours = total_seconds / 3600
            return Decimal(str(round(hours, 4)))
        
        # Se for número (int/float/Decimal) - Excel retorna dias como decimal
        if isinstance(value, (int, float, Decimal)):
            # Excel armazena tempo como fração de dia (ex: 0.5 = 12 horas)
            # Se for menor que 1, provavelmente é fração de dia
            num_value = float(value)
            if 0 < num_value < 1:
                # Converter dias para horas
                hours = num_value * 24
                return Decimal(str(round(hours, 2)))
            else:
                # Se for >= 1, pode ser dias (pouco provável) ou já estar em horas
                # Para segurança, assumir que valores grandes já estão em horas
                return Decimal(str(round(num_value, 2)))
        
        # Se for string, tentar parsear
        if isinstance(value, str):
            value = value.strip()
            
            if not value:
                return Decimal('0')
            
            try:
                # Tentar converter diretamente para número
                num = float(value)
                # Se for decimal pequeno, converter de dias para horas
                if 0 < num < 1:
                    return Decimal(str(round(num * 24, 2)))
                return Decimal(str(round(num, 2)))
            except:
                pass
            
            # Parsear formato "HH:MM:SS" ou "HHHH:MM:SS" (formato de horas totais)
            if ':' in value and 'day' not in value.lower():
                try:
                    time_parts = value.split(':')
                    hours = int(time_parts[0])
                    minutes = int(time_parts[1]) if len(time_parts) > 1 else 0
                    seconds = int(time_parts[2]) if len(time_parts) > 2 else 0
                    
                    total_hours = hours + (minutes / 60) + (seconds / 3600)
                    return Decimal(str(round(total_hours, 2)))
                except:
                    pass
            
            # Parsear formato "X days, HH:MM:SS"
            if 'day' in value.lower():
                try:
                    total_hours = 0
                    
                    # Extrair dias
                    parts = value.lower().split('day')
                    days = int(parts[0].strip())
                    total_hours += days * 24
                    
                    # Extrair horas:minutos:segundos da parte restante
                    time_str = parts[1].replace('s,', '').replace(',', '').strip()
                    if ':' in time_str:
                        time_parts = time_str.split(':')
                        total_hours += int(time_parts[0])
                        total_hours += int(time_parts[1]) / 60 if len(time_parts) > 1 else 0
                        total_hours += int(time_parts[2]) / 3600 if len(time_parts) > 2 else 0
                    
                    return Decimal(str(round(total_hours, 2)))
                except:
                    pass
        
        return Decimal('0')
    
    def _create_processing_log(
        self,
        period: TimecardPeriod,
        filename: str,
        results: Dict,
        processing_time: float
    ):
        """Cria log de processamento"""
        try:
            import os
            
            status = 'completed' if results['error_rows'] == 0 else 'partial'
            
            log = TimecardProcessingLog(
                period_id=period.id,
                filename=os.path.basename(filename),
                file_size=os.path.getsize(filename) if os.path.exists(filename) else None,
                total_rows=results['total_rows'],
                processed_rows=results['processed_rows'],
                error_rows=results['error_rows'],
                status=status,
                processing_summary={
                    'warnings_count': len(self.warnings),
                    'errors_count': len(self.errors)
                },
                processed_by=self.user_id,
                processing_time=Decimal(str(round(processing_time, 2)))
            )
            
            self.db.add(log)
            logger.info(f"Log de processamento criado: {filename}")
            
        except Exception as e:
            logger.error(f"Erro ao criar log de processamento: {e}")
