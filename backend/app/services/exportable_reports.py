"""Servicos para relatorios exportaveis (XLSX)."""

from __future__ import annotations

from collections import defaultdict
import calendar
import ast
import json
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.models.employee import Employee
from app.models.payroll import BenefitsData, BenefitsPeriod, PayrollData, PayrollPeriod
from app.utils.parsers import normalize_cpf, parse_br_number


CURRENCY_COLUMNS = {
    "Salario Base",
    "Proventos",
    "Descontos",
    "Adiantamento",
    "Liquido",
    "Mobilidade",
    "Vale Alimentacao",
    "Vale Refeicao",
    "Saldo Livre",
}

DEFAULT_INSS_PATRONAL_RATE = 0.20
DEFAULT_FGTS_RATE = 0.08


def _as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    return float(parse_br_number(value))


def _parse_json_like(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple, set)):
        return value
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("{") or text.startswith("["):
            try:
                return json.loads(text)
            except Exception:
                try:
                    return ast.literal_eval(text)
                except Exception:
                    return value
    return value


def _as_dict_payload(payload: Any) -> Dict[str, Any]:
    parsed = _parse_json_like(payload)
    return parsed if isinstance(parsed, dict) else {}


def _get_first_number(payload: Optional[Dict[str, Any]], keys: Iterable[str]) -> float:
    payload_dict = _as_dict_payload(payload)
    if not payload_dict:
        return 0.0
    for key in keys:
        if key in payload_dict and payload_dict[key] not in (None, ""):
            return _as_float(payload_dict[key])
    return 0.0


def _normalize_key(value: Any) -> str:
    import unicodedata

    text = str(value or "")
    if "\\u" in text:
        try:
            text = bytes(text, "utf-8").decode("unicode_escape")
        except Exception:
            pass

    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = "".join(ch if ch.isalnum() else " " for ch in text)
    return " ".join(text.split())


def _get_first_number_fuzzy(payload: Optional[Dict[str, Any]], aliases: Iterable[str]) -> float:
    payload_dict = _as_dict_payload(payload)
    if not payload_dict:
        return 0.0

    normalized_aliases = {_normalize_key(alias) for alias in aliases}
    for key, value in payload_dict.items():
        if _normalize_key(key) in normalized_aliases and value not in (None, ""):
            return _as_float(value)
    return 0.0


def _sum_numeric_payload(payload: Any) -> float:
    payload = _parse_json_like(payload)
    if payload is None:
        return 0.0
    if isinstance(payload, dict):
        return sum(_sum_numeric_payload(value) for value in payload.values())
    if isinstance(payload, (list, tuple, set)):
        return sum(_sum_numeric_payload(value) for value in payload)
    return _as_float(payload)


def _sum_numeric_payload_excluding_aliases(payload: Any, excluded_aliases: Iterable[str]) -> float:
    payload = _parse_json_like(payload)
    excluded = {_normalize_key(alias) for alias in excluded_aliases}

    if payload is None:
        return 0.0
    if isinstance(payload, dict):
        total = 0.0
        for key, value in payload.items():
            if _normalize_key(key) in excluded:
                continue
            total += _sum_numeric_payload(value)
        return total
    if isinstance(payload, (list, tuple, set)):
        return sum(_sum_numeric_payload(item) for item in payload)
    return _as_float(payload)


def _sum_numbers_by_aliases(payload: Optional[Dict[str, Any]], aliases: Iterable[str]) -> float:
    payload_dict = _as_dict_payload(payload)
    if not payload_dict:
        return 0.0

    normalized_aliases = {_normalize_key(alias) for alias in aliases}
    total = 0.0
    for key, value in payload_dict.items():
        if _normalize_key(key) in normalized_aliases and value not in (None, ""):
            total += _as_float(value)
    return total


def _latest_by_employee(rows: List[Any]) -> Dict[int, Any]:
    latest: Dict[int, Any] = {}
    for row in rows:
        if row.employee_id not in latest:
            latest[row.employee_id] = row
    return latest


def _active_employees_for_month(
    session: Session,
    company: str,
    year: int,
    month: int,
    department: Optional[str] = None,
    employee_id: Optional[int] = None,
) -> List[Employee]:
    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])

    query = (
        session.query(Employee)
        .filter(
            or_(
                Employee.company_code == company,
                Employee.unique_id.like(f"{company}%"),
                Employee.absolute_id.like(f"{company}%"),
            )
        )
        .filter(or_(Employee.admission_date.is_(None), Employee.admission_date <= month_end))
        .filter(or_(Employee.termination_date.is_(None), Employee.termination_date >= month_start))
    )

    if department:
        query = query.filter(Employee.department == department)

    if employee_id is not None:
        query = query.filter(Employee.id == employee_id)

    return query.order_by(Employee.name.asc()).all()

def _classify_payroll_period_type(period_name: Optional[str]) -> str:
    name = _normalize_key(period_name)
    if "13" in name and "adiant" in name:
        return "13_adiantamento"
    if "13" in name:
        return "13_integral"
    if "complement" in name:
        return "complementar"
    if "adiant" in name:
        return "adiantamento_salario"
    return "mensal"


def _latest_payroll_by_employee(
    session: Session,
    company: str,
    year: int,
    month: int,
    payroll_type: str = "mensal",
) -> Dict[int, PayrollData]:
    rows = (
        session.query(PayrollData, PayrollPeriod)
        .join(PayrollPeriod, PayrollPeriod.id == PayrollData.period_id)
        .filter(
            PayrollPeriod.company == company,
            PayrollPeriod.year == year,
            PayrollPeriod.month == month,
        )
        .order_by(
            PayrollData.employee_id.asc(),
            PayrollData.updated_at.desc().nullslast(),
            PayrollData.id.desc(),
        )
        .all()
    )

    filtered_rows: List[PayrollData] = []
    for payroll_data, period in rows:
        detected_type = _classify_payroll_period_type(period.period_name)
        if payroll_type != "all" and detected_type != payroll_type:
            continue
        filtered_rows.append(payroll_data)

    return _latest_by_employee(filtered_rows)


def _historical_monthly_payroll_rows(
    session: Session,
    company: str,
    year: int,
    month: int,
) -> Dict[int, List[PayrollData]]:
    rows = (
        session.query(PayrollData, PayrollPeriod)
        .join(PayrollPeriod, PayrollPeriod.id == PayrollData.period_id)
        .filter(
            PayrollPeriod.company == company,
            or_(
                PayrollPeriod.year < year,
                (PayrollPeriod.year == year) & (PayrollPeriod.month <= month),
            ),
        )
        .order_by(
            PayrollData.employee_id.asc(),
            PayrollPeriod.year.desc(),
            PayrollPeriod.month.desc(),
            PayrollData.updated_at.desc().nullslast(),
            PayrollData.id.desc(),
        )
        .all()
    )

    grouped: Dict[int, Dict[Tuple[int, int], PayrollData]] = defaultdict(dict)
    for payroll_data, period in rows:
        period_type = _classify_payroll_period_type(period.period_name)
        if period_type != "mensal":
            continue
        period_key = (period.year, period.month)
        if period_key not in grouped[payroll_data.employee_id]:
            grouped[payroll_data.employee_id][period_key] = payroll_data

    history: Dict[int, List[PayrollData]] = {}
    for employee_id, period_map in grouped.items():
        sorted_periods = sorted(period_map.keys(), reverse=True)
        history[employee_id] = [period_map[key] for key in sorted_periods[:12]]
    return history


def _latest_benefits_by_employee(session: Session, company: str, year: int, month: int) -> Dict[int, BenefitsData]:
    rows = (
        session.query(BenefitsData)
        .join(BenefitsPeriod, BenefitsPeriod.id == BenefitsData.period_id)
        .filter(
            BenefitsPeriod.company == company,
            BenefitsPeriod.year == year,
            BenefitsPeriod.month == month,
        )
        .order_by(
            BenefitsData.employee_id.asc(),
            BenefitsData.updated_at.desc().nullslast(),
            BenefitsData.id.desc(),
        )
        .all()
    )
    return _latest_by_employee(rows)


def _last_day_of_month(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def _completed_months_between(start_date: date, end_date: date) -> int:
    if start_date > end_date:
        return 0

    months = (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month)
    if end_date.day < start_date.day:
        months -= 1
    return max(months, 0)


def _annual_cycle_months(reference_date: date, admission_date: Optional[date]) -> int:
    if admission_date is None or admission_date > reference_date:
        return 0

    anchor_day = min(admission_date.day, _last_day_of_month(reference_date.year, admission_date.month))
    anniversary = date(reference_date.year, admission_date.month, anchor_day)
    if anniversary > reference_date:
        prev_year = reference_date.year - 1
        anchor_day = min(admission_date.day, _last_day_of_month(prev_year, admission_date.month))
        anniversary = date(prev_year, admission_date.month, anchor_day)

    return min(_completed_months_between(anniversary, reference_date), 11)


def _extract_financial_components(payroll: Optional[PayrollData]) -> Dict[str, float]:
    adiantamento_aliases = [
        "ADIANTAMENTO",
        "Adiantamento",
        "Adiantamento Salarial",
        "ADIANTAMENTO_SALARIAL",
        "Adiantamento 13",
        "Adiantamento_13",
        "13 Adiantamento",
        "13o Adiantamento",
    ]

    payroll_additional = _as_dict_payload(payroll.additional_data) if payroll else {}
    payroll_earnings = _as_dict_payload(payroll.earnings_data) if payroll else {}
    payroll_deductions = _as_dict_payload(payroll.deductions_data) if payroll else {}

    salario_base = _get_first_number_fuzzy(
        payroll_additional,
        ["Salário Mensal", "Salario Mensal", "Valor Salário", "Valor Salario", "Salário Base", "Salario Base"],
    )
    if not salario_base:
        salario_base = _get_first_number_fuzzy(
            payroll_earnings,
            ["SALARIO_BASE", "Salário Base", "Salario Base", "SALARIO MENSAL", "Salário Mensal"],
        )
    if not salario_base and payroll is not None:
        salario_base = _as_float(payroll.gross_salary)

    proventos = _get_first_number_fuzzy(
        payroll_additional,
        [
            "Total de Proventos",
            "TOTAL_PROVENTOS",
            "Total Proventos",
            "Total Provento",
            "Proventos",
            "Total Vencimentos",
        ],
    )
    if not proventos and payroll is not None:
        proventos = _as_float(payroll.gross_salary)
    if not proventos:
        proventos = _sum_numeric_payload(payroll_earnings)

    descontos = _get_first_number_fuzzy(
        payroll_additional,
        [
            "Total de Descontos",
            "TOTAL_DESCONTOS",
            "Total Descontos",
            "Descontos",
            "Total de Deducoes",
        ],
    )
    if not descontos:
        descontos = _sum_numeric_payload(payroll_deductions)

    adiantamento = _get_first_number_fuzzy(payroll_deductions, adiantamento_aliases)
    if not adiantamento:
        adiantamento = _get_first_number_fuzzy(payroll_additional, adiantamento_aliases)

    descontos_sem_adiantamento = 0.0
    if descontos > 0:
        descontos_sem_adiantamento = max(descontos - adiantamento, 0.0)
    else:
        descontos_sem_adiantamento = _sum_numeric_payload_excluding_aliases(payroll_deductions, adiantamento_aliases)

    liquido = _get_first_number_fuzzy(
        payroll_additional,
        ["Liquido de Calculo", "Líquido de Cálculo", "LIQ_A_RECEBER", "Liquido", "Liquido a Receber"],
    )
    if not liquido and payroll is not None:
        liquido = _as_float(payroll.net_salary)

    base_inss_patronal = _get_first_number_fuzzy(
        payroll_additional,
        [
            "Base INSS Normal - Empresa",
            "Base INSS Normal Empresa",
            "BASE_INSS_NORMAL_EMPRESA",
            "Base INSS Empresa",
        ],
    )
    if not base_inss_patronal:
        base_inss_patronal = _get_first_number_fuzzy(
            payroll_earnings,
            [
                "Base INSS Normal - Empresa",
                "Base INSS Normal Empresa",
                "BASE_INSS_NORMAL_EMPRESA",
                "Base INSS Empresa",
            ],
        )
    if not base_inss_patronal:
        base_inss_patronal = salario_base

    fgts_folha = _sum_numbers_by_aliases(
        payroll_deductions,
        [
            "FGTS",
        ],
    )

    return {
        "salario_base": salario_base,
        "proventos": proventos,
        "descontos": descontos_sem_adiantamento,
        "adiantamento": adiantamento,
        "liquido": liquido,
        "base_inss_patronal": base_inss_patronal,
        "fgts_folha": fgts_folha,
    }


def _write_generic_xlsx(
    title: str,
    sheet_name: str,
    rows: List[Dict[str, Any]],
    columns: List[str],
    currency_columns: Iterable[str],
    date_columns: Iterable[str],
    summary_pairs: List[Tuple[str, Any]],
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    zebra_fill = PatternFill("solid", fgColor="F5FAFF")
    thin = Side(border_style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = center

    header_row = 3
    for col_idx, name in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    currency_set = set(currency_columns)
    date_set = set(date_columns)
    for row_idx, row_data in enumerate(rows, start=header_row + 1):
        row_fill = zebra_fill if row_idx % 2 == 0 else None
        for col_idx, name in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=row_data.get(name))
            cell.border = border
            if row_fill:
                cell.fill = row_fill

            if name in currency_set:
                cell.number_format = 'R$ #,##0.00'
            elif name in date_set and cell.value:
                cell.number_format = 'DD/MM/YYYY'

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + max(1, len(rows))}"

    for idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for row_data in rows[:500]:
            max_len = max(max_len, len(str(row_data.get(col_name, ""))))
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(12, max_len + 2), 36)

    summary = workbook.create_sheet("Resumo")
    summary["A1"] = "Resumo da exportacao"
    summary["A1"].font = Font(bold=True, size=12, color="1F4E79")
    line = 3
    for key, value in summary_pairs:
        summary[f"A{line}"] = key
        summary[f"B{line}"] = value
        line += 1
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 42

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.getvalue()


def _build_export_rows(
    employees: List[Employee],
    payroll_by_employee: Dict[int, PayrollData],
    benefits_by_employee: Dict[int, BenefitsData],
    company: str,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    for employee in employees:
        payroll = payroll_by_employee.get(employee.id)
        benefits = benefits_by_employee.get(employee.id)
        financial = _extract_financial_components(payroll)

        rows.append(
            {
                "Nome": employee.name,
                "Cargo": employee.position or "",
                "Setor": employee.department or "",
                "CPF": normalize_cpf(employee.cpf) if employee.cpf else "",
                "Matricula": employee.unique_id or employee.registration_number or "",
                "Absolute ID": employee.absolute_id or "",
                "Empresa": employee.company_code or company,
                "Situacao": employee.employment_status or "",
                "Admissao": employee.admission_date,
                "Demissao": employee.termination_date,
                "Salario Base": financial["salario_base"],
                "Proventos": financial["proventos"],
                "Descontos": financial["descontos"],
                "Adiantamento": financial["adiantamento"],
                "Liquido": financial["liquido"],
                "Mobilidade": _as_float(benefits.mobilidade) if benefits else 0.0,
                "Vale Alimentacao": _as_float(benefits.alimentacao) if benefits else 0.0,
                "Vale Refeicao": _as_float(benefits.refeicao) if benefits else 0.0,
                "Saldo Livre": _as_float(benefits.livre) if benefits else 0.0,
            }
        )

    return rows


def _write_xlsx(rows: List[Dict[str, Any]], company: str, year: int, month: int) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{company} {year}-{month:02d}"

    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    zebra_fill = PatternFill("solid", fgColor="F5FAFF")
    thin = Side(border_style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    columns = [
        "Nome",
        "Cargo",
        "Setor",
        "CPF",
        "Matricula",
        "Absolute ID",
        "Empresa",
        "Situacao",
        "Admissao",
        "Demissao",
        "Salario Base",
        "Proventos",
        "Descontos",
        "Adiantamento",
        "Liquido",
        "Mobilidade",
        "Vale Alimentacao",
        "Vale Refeicao",
        "Saldo Livre",
    ]

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = sheet.cell(row=1, column=1, value=f"Relatorio Estrategico - Infraestrutura {company} - {month:02d}/{year}")
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = center

    header_row = 3
    for col_idx, name in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    for row_idx, row_data in enumerate(rows, start=header_row + 1):
        row_fill = zebra_fill if row_idx % 2 == 0 else None
        for col_idx, name in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=row_data.get(name))
            cell.border = border
            if row_fill:
                cell.fill = row_fill

            if name in CURRENCY_COLUMNS:
                cell.number_format = 'R$ #,##0.00'
            elif name in {"Admissao", "Demissao"} and cell.value:
                cell.number_format = 'DD/MM/YYYY'

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + max(1, len(rows))}"

    widths = {
        "A": 28,
        "B": 22,
        "C": 18,
        "D": 16,
        "E": 15,
        "F": 24,
        "G": 10,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 14,
        "O": 14,
        "P": 16,
        "Q": 16,
        "R": 14,
        "S": 14,
    }
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width

    summary = workbook.create_sheet("Resumo")
    summary["A1"] = "Resumo da exportacao"
    summary["A1"].font = Font(bold=True, size=12, color="1F4E79")
    summary["A3"] = "Empresa"
    summary["B3"] = company
    summary["A4"] = "Competencia"
    summary["B4"] = f"{year}-{month:02d}"
    summary["A5"] = "Colaboradores"
    summary["B5"] = len(rows)
    summary["A6"] = "Gerado pelo"
    summary["B6"] = "Relatorios Exportaveis"
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 26

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.getvalue()


def build_infra_analytics_xlsx(
    session: Session,
    year: int,
    month: int,
    company: str = "0059",
    payroll_type: str = "mensal",
    department: Optional[str] = None,
    employee_id: Optional[int] = None,
) -> Tuple[bytes, int, str]:
    """Gera o xlsx de relatorio estrategico de infraestrutura e retorna bytes, total de linhas e nome sugerido."""
    employees = _active_employees_for_month(
        session,
        company=company,
        year=year,
        month=month,
        department=department,
        employee_id=employee_id,
    )
    payroll = _latest_payroll_by_employee(
        session,
        company=company,
        year=year,
        month=month,
        payroll_type=payroll_type,
    )
    benefits = _latest_benefits_by_employee(session, company=company, year=year, month=month)

    rows = _build_export_rows(employees, payroll, benefits, company=company)
    xlsx_bytes = _write_xlsx(rows, company=company, year=year, month=month)
    suffix = "" if payroll_type == "mensal" else f"_{payroll_type}"
    filename = f"relatorio_estrategico_{company}_{year}-{month:02d}{suffix}.xlsx"
    return xlsx_bytes, len(rows), filename


def build_monthly_provisions_xlsx(
    session: Session,
    year: int,
    month: int,
    company: str = "0059",
    payroll_type: str = "mensal",
    department: Optional[str] = None,
    employee_id: Optional[int] = None,
    inss_patronal_rate: float = DEFAULT_INSS_PATRONAL_RATE,
    fgts_rate: float = DEFAULT_FGTS_RATE,
) -> Tuple[bytes, int, str]:
    employees = _active_employees_for_month(
        session,
        company=company,
        year=year,
        month=month,
        department=department,
        employee_id=employee_id,
    )
    payroll = _latest_payroll_by_employee(
        session,
        company=company,
        year=year,
        month=month,
        payroll_type=payroll_type,
    )
    benefits = _latest_benefits_by_employee(session, company=company, year=year, month=month)

    rows: List[Dict[str, Any]] = []
    for employee in employees:
        payroll_row = payroll.get(employee.id)
        benefits_row = benefits.get(employee.id)
        financial = _extract_financial_components(payroll_row)

        salario_base = financial["salario_base"]
        proventos = financial["proventos"]
        liquido = financial["liquido"]
        base_inss_patronal = financial["base_inss_patronal"]
        fgts_folha = financial["fgts_folha"]

        beneficios_total = (
            (_as_float(benefits_row.mobilidade) if benefits_row else 0.0)
            + (_as_float(benefits_row.alimentacao) if benefits_row else 0.0)
            + (_as_float(benefits_row.refeicao) if benefits_row else 0.0)
            + (_as_float(benefits_row.livre) if benefits_row else 0.0)
        )

        provisao_ferias = salario_base / 12 if salario_base else 0.0
        provisao_terco_ferias = provisao_ferias / 3 if provisao_ferias else 0.0
        provisao_decimo = salario_base / 12 if salario_base else 0.0

        inss_patronal_mensal = base_inss_patronal * inss_patronal_rate if base_inss_patronal else 0.0
        fgts_provisionado = salario_base * fgts_rate if salario_base else 0.0
        fgts_mensal = fgts_folha if fgts_folha > 0 else fgts_provisionado
        inss_provisoes = (provisao_ferias + provisao_terco_ferias + provisao_decimo) * inss_patronal_rate

        total_provisoes = provisao_ferias + provisao_terco_ferias + provisao_decimo + inss_provisoes
        custo_mensal_estimado = proventos + beneficios_total + inss_patronal_mensal + fgts_mensal + total_provisoes

        rows.append(
            {
                "Nome": employee.name,
                "Setor": employee.department or "",
                "Cargo": employee.position or "",
                "Empresa": employee.company_code or company,
                "Salario Base": salario_base,
                "Base INSS Patronal": base_inss_patronal,
                "Proventos": proventos,
                "Liquido": liquido,
                "Beneficios": beneficios_total,
                "INSS Patronal (Mensal)": inss_patronal_mensal,
                "FGTS (Mensal)": fgts_mensal,
                "FGTS Provisionado (8%)": fgts_provisionado,
                "Provisao 1/12 Ferias": provisao_ferias,
                "Provisao 1/3 Ferias": provisao_terco_ferias,
                "Provisao 1/12 13": provisao_decimo,
                "INSS sobre Provisoes": inss_provisoes,
                "Total Provisoes": total_provisoes,
                "Custo Mensal Estimado": custo_mensal_estimado,
            }
        )

    summary_pairs = [
        ("Empresa", company),
        ("Competencia", f"{year}-{month:02d}"),
        ("Tipo de folha", payroll_type),
        ("Colaboradores", len(rows)),
        ("Aliquota INSS Patronal", f"{inss_patronal_rate * 100:.2f}%"),
        ("Aliquota FGTS", f"{fgts_rate * 100:.2f}%"),
        ("Regra", "MVP de provisoes mensais para caixa"),
    ]

    columns = [
        "Nome",
        "Setor",
        "Cargo",
        "Empresa",
        "Salario Base",
        "Base INSS Patronal",
        "Proventos",
        "Liquido",
        "Beneficios",
        "INSS Patronal (Mensal)",
        "FGTS (Mensal)",
        "FGTS Provisionado (8%)",
        "Provisao 1/12 Ferias",
        "Provisao 1/3 Ferias",
        "Provisao 1/12 13",
        "INSS sobre Provisoes",
        "Total Provisoes",
        "Custo Mensal Estimado",
    ]

    currency_columns = {
        "Salario Base",
        "Base INSS Patronal",
        "Proventos",
        "Liquido",
        "Beneficios",
        "INSS Patronal (Mensal)",
        "FGTS (Mensal)",
        "FGTS Provisionado (8%)",
        "Provisao 1/12 Ferias",
        "Provisao 1/3 Ferias",
        "Provisao 1/12 13",
        "INSS sobre Provisoes",
        "Total Provisoes",
        "Custo Mensal Estimado",
    }

    xlsx_bytes = _write_generic_xlsx(
        title=f"Relatorio de Provisoes Mensais - {company} - {month:02d}/{year}",
        sheet_name="Provisoes",
        rows=rows,
        columns=columns,
        currency_columns=currency_columns,
        date_columns=set(),
        summary_pairs=summary_pairs,
    )
    filename = f"relatorio_provisoes_{company}_{year}-{month:02d}.xlsx"
    return xlsx_bytes, len(rows), filename


def build_termination_simulation_xlsx(
    session: Session,
    year: int,
    month: int,
    company: str = "0059",
    payroll_type: str = "mensal",
    department: Optional[str] = None,
    employee_id: Optional[int] = None,
    scenario: str = "sem_justa_causa",
    termination_day: int = 15,
    fgts_rate: float = DEFAULT_FGTS_RATE,
) -> Tuple[bytes, int, str]:
    employees = _active_employees_for_month(
        session,
        company=company,
        year=year,
        month=month,
        department=department,
        employee_id=employee_id,
    )
    payroll = _latest_payroll_by_employee(
        session,
        company=company,
        year=year,
        month=month,
        payroll_type=payroll_type,
    )

    day = min(max(1, termination_day), _last_day_of_month(year, month))
    reference_date = date(year, month, day)

    rows: List[Dict[str, Any]] = []
    for employee in employees:
        payroll_row = payroll.get(employee.id)
        financial = _extract_financial_components(payroll_row)

        salario_base = financial["salario_base"]
        adiantamento = financial["adiantamento"]
        admission_date = employee.admission_date

        months_service = _completed_months_between(admission_date, reference_date) if admission_date else 0
        years_service = months_service // 12
        annual_cycle_months = _annual_cycle_months(reference_date, admission_date)
        months_13 = max(reference_date.month - (1 if reference_date.day < 15 else 0), 0)

        saldo_salario = (salario_base / 30.0) * day if salario_base else 0.0
        decimo_proporcional = (salario_base / 12.0) * months_13 if salario_base else 0.0
        ferias_proporcionais = (salario_base / 12.0) * annual_cycle_months if salario_base else 0.0
        terco_ferias = ferias_proporcionais / 3.0 if ferias_proporcionais else 0.0

        aviso_dias = 30
        if years_service > 1:
            aviso_dias = min(90, 30 + ((years_service - 1) * 3))
        aviso_previo = (salario_base / 30.0) * aviso_dias if salario_base else 0.0

        fgts_saldo_estimado = salario_base * fgts_rate * months_service if salario_base else 0.0
        multa_fgts_40 = fgts_saldo_estimado * 0.40

        if scenario in {"pedido_demissao", "termino_contrato"}:
            aviso_previo = 0.0
            multa_fgts_40 = 0.0

        total_bruto = saldo_salario + decimo_proporcional + ferias_proporcionais + terco_ferias + aviso_previo + multa_fgts_40
        total_liquido = max(total_bruto - adiantamento, 0.0)

        rows.append(
            {
                "Nome": employee.name,
                "Setor": employee.department or "",
                "Cargo": employee.position or "",
                "Empresa": employee.company_code or company,
                "Admissao": admission_date,
                "Referencia Simulacao": reference_date,
                "Cenario": scenario,
                "Salario Base": salario_base,
                "Meses no Contrato": months_service,
                "Meses Proporcionais (Ferias)": annual_cycle_months,
                "Meses Proporcionais (13)": months_13,
                "Saldo de Salario": saldo_salario,
                "13 Proporcional": decimo_proporcional,
                "Ferias Proporcionais": ferias_proporcionais,
                "1/3 Ferias": terco_ferias,
                "Aviso Previo": aviso_previo,
                "Multa FGTS 40%": multa_fgts_40,
                "Adiantamento a Descontar": adiantamento,
                "Total Rescisao Bruta": total_bruto,
                "Total Rescisao Liquida": total_liquido,
            }
        )

    summary_pairs = [
        ("Empresa", company),
        ("Competencia", f"{year}-{month:02d}"),
        ("Tipo de folha", payroll_type),
        ("Cenario", scenario),
        ("Dia de referencia", day),
        ("Aliquota FGTS", f"{fgts_rate * 100:.2f}%"),
        ("Observacao", "MVP com estimativa automatica para planejamento financeiro"),
    ]

    columns = [
        "Nome",
        "Setor",
        "Cargo",
        "Empresa",
        "Admissao",
        "Referencia Simulacao",
        "Cenario",
        "Salario Base",
        "Meses no Contrato",
        "Meses Proporcionais (Ferias)",
        "Meses Proporcionais (13)",
        "Saldo de Salario",
        "13 Proporcional",
        "Ferias Proporcionais",
        "1/3 Ferias",
        "Aviso Previo",
        "Multa FGTS 40%",
        "Adiantamento a Descontar",
        "Total Rescisao Bruta",
        "Total Rescisao Liquida",
    ]
    currency_columns = {
        "Salario Base",
        "Saldo de Salario",
        "13 Proporcional",
        "Ferias Proporcionais",
        "1/3 Ferias",
        "Aviso Previo",
        "Multa FGTS 40%",
        "Adiantamento a Descontar",
        "Total Rescisao Bruta",
        "Total Rescisao Liquida",
    }

    xlsx_bytes = _write_generic_xlsx(
        title=f"Simulacao de Rescisao - {company} - {month:02d}/{year}",
        sheet_name="Rescisoes",
        rows=rows,
        columns=columns,
        currency_columns=currency_columns,
        date_columns={"Admissao", "Referencia Simulacao"},
        summary_pairs=summary_pairs,
    )
    filename = f"simulacao_rescisoes_{company}_{year}-{month:02d}_{scenario}.xlsx"
    return xlsx_bytes, len(rows), filename


def build_fund_personnel_cost_xlsx(
    session: Session,
    year: int,
    month: int,
    company: str = "0059",
    payroll_type: str = "mensal",
    department: Optional[str] = None,
    employee_id: Optional[int] = None,
    inss_patronal_rate: float = DEFAULT_INSS_PATRONAL_RATE,
) -> Tuple[bytes, int, str]:
    employees = _active_employees_for_month(
        session,
        company=company,
        year=year,
        month=month,
        department=department,
        employee_id=employee_id,
    )
    payroll = _latest_payroll_by_employee(
        session,
        company=company,
        year=year,
        month=month,
        payroll_type=payroll_type,
    )
    benefits = _latest_benefits_by_employee(session, company=company, year=year, month=month)
    payroll_history = _historical_monthly_payroll_rows(session, company=company, year=year, month=month)

    rows: List[Dict[str, Any]] = []
    for employee in employees:
        payroll_row = payroll.get(employee.id)
        benefits_row = benefits.get(employee.id)
        financial = _extract_financial_components(payroll_row)

        earnings_payload = _as_dict_payload(payroll_row.earnings_data) if payroll_row else {}
        deductions_payload = _as_dict_payload(payroll_row.deductions_data) if payroll_row else {}
        additional_payload = _as_dict_payload(payroll_row.additional_data) if payroll_row else {}

        salario_base = financial["salario_base"]
        proventos = financial["proventos"]
        liquido = financial["liquido"]
        descontos = financial["descontos"]
        base_inss_patronal = financial["base_inss_patronal"]

        he_50 = _sum_numbers_by_aliases(earnings_payload, ["HE_50_DIURNAS"])
        he_100 = _sum_numbers_by_aliases(earnings_payload, ["HE_100_DIURNAS"])
        he_not_50 = _sum_numbers_by_aliases(earnings_payload, ["HE_50_NOTURNAS"])
        he_not_100 = _sum_numbers_by_aliases(earnings_payload, ["HE_100_NOTURNAS"])
        he_total = he_50 + he_100 + he_not_50 + he_not_100
        dsr = _sum_numbers_by_aliases(earnings_payload, ["DSR_HE_DIURNAS", "DSR_HE_NOTURNAS"])
        total_he = he_total

        periculosidade = _sum_numbers_by_aliases(earnings_payload, ["PERICULOSIDADE"])
        insalubridade = _sum_numbers_by_aliases(earnings_payload, ["INSALUBRIDADE"])

        inss_patronal = base_inss_patronal * inss_patronal_rate if base_inss_patronal else 0.0
        fgts = _sum_numbers_by_aliases(deductions_payload, ["FGTS"])
        ir = _sum_numbers_by_aliases(deductions_payload, ["IRRF"])
        inss = _sum_numbers_by_aliases(deductions_payload, ["INSS"])

        beneficio_mobilidade = _as_float(benefits_row.mobilidade) if benefits_row else 0.0
        beneficio_alimentacao = _as_float(benefits_row.alimentacao) if benefits_row else 0.0
        beneficio_refeicao = _as_float(benefits_row.refeicao) if benefits_row else 0.0
        beneficio_livre = _as_float(benefits_row.livre) if benefits_row else 0.0
        total_beneficios = beneficio_mobilidade + beneficio_alimentacao + beneficio_refeicao + beneficio_livre

        history_rows = payroll_history.get(employee.id, [])
        proventos_history = []
        for hrow in history_rows:
            h_financial = _extract_financial_components(hrow)
            if h_financial["proventos"] > 0:
                proventos_history.append(h_financial["proventos"])

        remuneracao_referencia = proventos if proventos > 0 else salario_base
        meses_historico = min(len(proventos_history), 12)
        soma_historico = sum(proventos_history[:12])
        meses_faltantes = max(12 - meses_historico, 0)
        soma_12_ajustada = soma_historico + (remuneracao_referencia * meses_faltantes)

        media_proventos_12m = soma_12_ajustada / 12 if soma_12_ajustada else 0.0
        # Provisao mensal: 1/12 da media de 12 meses.
        ferias_mes = media_proventos_12m / 12 if media_proventos_12m else 0.0
        decimo_mes = media_proventos_12m / 12 if media_proventos_12m else 0.0
        # Provisao mensal do terco constitucional sobre a provisao de ferias.
        terco_ferias_mes = ferias_mes / 3 if ferias_mes else 0.0
        total_provisoes = ferias_mes + terco_ferias_mes + decimo_mes

        total_bruto = proventos + total_beneficios + ferias_mes + decimo_mes
        total_bruto_provisoes = total_bruto + total_provisoes
        total_liquido_beneficio = liquido + total_beneficios

        horas_mensais = _get_first_number_fuzzy(
            additional_payload,
            ["Horas Normais Diurnas", "Horas Normais", "HORAS_NORMAIS"],
        )
        if horas_mensais <= 0:
            horas_mensais = 220.0

        custo_hora_bruta = total_bruto / horas_mensais if horas_mensais else 0.0
        custo_hora_liquida = total_liquido_beneficio / horas_mensais if horas_mensais else 0.0

        rows.append(
            {
                "Funcionario": employee.name,
                "Setor": employee.department or "",
                "Funcao": employee.position or "",
                "Local de Trabalho": employee.work_location.name if getattr(employee, "work_location", None) else "",
                "Base": salario_base,
                "HE 50%": he_50,
                "HE 100%": he_100,
                "HE Not 50%": he_not_50,
                "HE Not 100%": he_not_100,
                "DSR": dsr,
                "Total HE": total_he,
                "Periculosidade": periculosidade,
                "Insalubridade": insalubridade,
                "Total Proventos": proventos,
                "INSS Patronal": inss_patronal,
                "FGTS": fgts,
                "IR": ir,
                "INSS": inss,
                "Total Descontos": descontos,
                "Total Líquido": liquido,
                "Mobilidade": beneficio_mobilidade,
                "Alimentação": beneficio_alimentacao,
                "Refeição": beneficio_refeicao,
                "Saldo Livre": beneficio_livre,
                "Total Benefícios": total_beneficios,
                "Férias - 1/12": ferias_mes,
                "1/3 Férias": terco_ferias_mes,
                "13º - 1/12": decimo_mes,
                "Total Provisões": total_provisoes,
                "Total Bruto": total_bruto,
                "Total Bruto + Provisões": total_bruto_provisoes,
                "Total Líquido + Benefícios": total_liquido_beneficio,
                "Custo Hora Bruta": custo_hora_bruta,
                "Custo Hora Líquida": custo_hora_liquida,
            }
        )

    columns = [
        "Funcionario",
        "Setor",
        "Funcao",
        "Local de Trabalho",
        "Base",
        "HE 50%",
        "HE 100%",
        "HE Not 50%",
        "HE Not 100%",
        "DSR",
        "Total HE",
        "Periculosidade",
        "Insalubridade",
        "Total Proventos",
        "INSS Patronal",
        "FGTS",
        "IR",
        "INSS",
        "Total Descontos",
        "Total Líquido",
        "Mobilidade",
        "Alimentação",
        "Refeição",
        "Saldo Livre",
        "Total Benefícios",
        "Férias - 1/12",
        "1/3 Férias",
        "13º - 1/12",
        "Total Provisões",
        "Total Bruto",
        "Total Bruto + Provisões",
        "Total Líquido + Benefícios",
        "Custo Hora Bruta",
        "Custo Hora Líquida",
    ]

    currency_columns = set(columns[4:])

    summary_pairs = [
        ("Titulo", "Relatorio de Custo de Pessoal para o Fundo"),
        ("Empresa", company),
        ("Competencia", f"{year}-{month:02d}"),
        ("Tipo de folha", payroll_type),
        ("Colaboradores", len(rows)),
        ("Aliquota INSS Patronal", f"{inss_patronal_rate * 100:.2f}%"),
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Custo Fundo"

    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    zebra_fill = PatternFill("solid", fgColor="F5FAFF")
    thin = Side(border_style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = sheet.cell(row=1, column=1, value="Relatorio de Custo de Pessoal para o Fundo")
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = center

    header_row = 3
    for col_idx, name in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    for row_idx, row_data in enumerate(rows, start=header_row + 1):
        row_fill = zebra_fill if row_idx % 2 == 0 else None
        for col_idx, name in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=row_data.get(name))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            if name in currency_columns:
                cell.number_format = 'R$ #,##0.00'

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + max(1, len(rows))}"
    for idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for row_data in rows[:500]:
            max_len = max(max_len, len(str(row_data.get(col_name, ""))))
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(12, max_len + 2), 34)

    summary = workbook.create_sheet("Resumo")
    summary["A1"] = "Resumo da exportacao"
    summary["A1"].font = Font(bold=True, size=12, color="1F4E79")
    line = 3
    for key, value in summary_pairs:
        summary[f"A{line}"] = key
        summary[f"B{line}"] = value
        line += 1
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 42

    totals = workbook.create_sheet("Totalizadores")
    totals["A1"] = "Bruto + Provisões por Setor"
    totals["A1"].font = Font(bold=True, size=12, color="1F4E79")
    totals["A3"] = "Setor"
    totals["B3"] = "Colaboradores"
    totals["C3"] = "Total Bruto + Provisões"
    for cell_ref in ("A3", "B3", "C3"):
        totals[cell_ref].font = Font(bold=True)

    setor_acc: Dict[str, Dict[str, float]] = defaultdict(lambda: {"count": 0.0, "total": 0.0})
    funcao_acc: Dict[str, Dict[str, float]] = defaultdict(lambda: {"count": 0.0, "total": 0.0})
    for row in rows:
        setor = str(row.get("Setor") or "Sem setor")
        funcao = str(row.get("Funcao") or "Sem funcao")
        total_bp = _as_float(row.get("Total Bruto + Provisões"))

        setor_acc[setor]["count"] += 1
        setor_acc[setor]["total"] += total_bp

        funcao_acc[funcao]["count"] += 1
        funcao_acc[funcao]["total"] += total_bp

    line = 4
    for setor in sorted(setor_acc.keys()):
        totals[f"A{line}"] = setor
        totals[f"B{line}"] = int(setor_acc[setor]["count"])
        totals[f"C{line}"] = setor_acc[setor]["total"]
        totals[f"C{line}"].number_format = 'R$ #,##0.00'
        line += 1

    line += 2
    totals[f"A{line}"] = "Bruto + Provisões por Função"
    totals[f"A{line}"].font = Font(bold=True, size=12, color="1F4E79")
    line += 2
    totals[f"A{line}"] = "Função"
    totals[f"B{line}"] = "Colaboradores"
    totals[f"C{line}"] = "Total Bruto + Provisões"
    for col in ("A", "B", "C"):
        totals[f"{col}{line}"].font = Font(bold=True)

    line += 1
    for funcao in sorted(funcao_acc.keys()):
        totals[f"A{line}"] = funcao
        totals[f"B{line}"] = int(funcao_acc[funcao]["count"])
        totals[f"C{line}"] = funcao_acc[funcao]["total"]
        totals[f"C{line}"].number_format = 'R$ #,##0.00'
        line += 1

    totals.column_dimensions["A"].width = 34
    totals.column_dimensions["B"].width = 16
    totals.column_dimensions["C"].width = 26

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    xlsx_bytes = stream.getvalue()
    filename = f"relatorio_custo_fundo_{company}_{year}-{month:02d}.xlsx"
    return xlsx_bytes, len(rows), filename
