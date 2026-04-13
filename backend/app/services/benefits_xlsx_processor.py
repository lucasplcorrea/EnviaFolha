"""
Serviço para processar arquivos XLSX de benefícios iFood
"""
import csv
import io
import logging
import calendar
import re
import openpyxl
from datetime import datetime, date
from collections import defaultdict
from decimal import Decimal
from typing import Dict, List, Optional, Tuple
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError

from app.models.company import Company
from app.models.employee import Employee
from app.models.payroll import BenefitsPeriod, BenefitsData, BenefitsProcessingLog, PayrollData, PayrollPeriod

logger = logging.getLogger(__name__)


class BenefitsXLSXProcessor:
    """Processador de arquivos XLSX de benefícios"""

    STRICT_SCHEMA_HEADERS = {
        'company_name': 'nome da empresa',
        'cnpj': 'cnpj',
        'recharge_id': 'id da recarga',
        'recharge_context': 'contexto da recarga',
        'recharge_month': 'mes da recarga',
        'cpf': 'cpf',
        'nome': 'nome do colaborador',
        'refeicao_pat': 'refeicao pat',
        'alimentacao_pat': 'alimentacao pat',
        'alimentacao_refeicao_nao_pat': 'alimentacao refeicao nao pat',
        'comer_ifood': 'comer no ifood',
        'mobilidade': 'mobilidade',
        'cultura': 'cultura',
        'home_office': 'home office',
        'educacao': 'educacao',
        'saude_bem_estar': 'saude bem estar',
        'farmacia': 'farmacia',
        'livre': 'livre',
    }
    
    def __init__(self, db_session: Session, user_id: Optional[int] = None):
        self.db = db_session
        self.user_id = user_id
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self._employee_name_map: Optional[Dict[str, Employee]] = None

    @staticmethod
    def _normalize_header(value: str) -> str:
        """Normaliza cabeçalhos para matching resiliente."""
        import re
        import unicodedata

        if value is None:
            return ""

        header = str(value).strip().lower()
        header = unicodedata.normalize("NFKD", header)
        header = "".join(ch for ch in header if not unicodedata.combining(ch))
        header = re.sub(r'[^a-z0-9 ]+', ' ', header)
        header = re.sub(r'\s+', ' ', header).strip()
        return header

    def _map_columns(self, headers: List[str]) -> Dict[str, int]:
        """Mapeia colunas do schema consolidado estrito."""
        normalized_index = {
            self._normalize_header(header): idx
            for idx, header in enumerate(headers)
        }

        mapping: Dict[str, int] = {}
        for target, normalized_header in self.STRICT_SCHEMA_HEADERS.items():
            normalized_expected = self._normalize_header(normalized_header)
            if normalized_expected in normalized_index:
                mapping[target] = normalized_index[normalized_expected]

        return mapping

    @staticmethod
    def _normalize_digits(value: str) -> str:
        return re.sub(r'\D', '', str(value or ''))

    @staticmethod
    def _normalize_company_name(value: str) -> str:
        return BenefitsXLSXProcessor._normalize_header(value)

    @staticmethod
    def _normalize_recharge_month(value) -> str:
        if value is None:
            return ""

        if isinstance(value, (datetime, date)):
            return f"{value.year:04d}-{value.month:02d}"

        text = str(value).strip()
        match = re.match(r'^(\d{4})[-/](\d{1,2})', text)
        if match:
            year = int(match.group(1))
            month = int(match.group(2))
            if 1 <= month <= 12:
                return f"{year:04d}-{month:02d}"
        return text

    def _resolve_company_constraints(self, company_code: str) -> Dict[str, set]:
        companies = self.db.query(Company).filter(Company.payroll_prefix == company_code).all()

        cnpjs = {
            self._normalize_digits(company.cnpj)
            for company in companies
            if company.cnpj
        }
        names = {
            self._normalize_company_name(company.name)
            for company in companies
            if company.name
        }
        trade_names = {
            self._normalize_company_name(company.trade_name)
            for company in companies
            if company.trade_name
        }

        all_names = {name for name in names.union(trade_names) if name}
        return {
            'cnpjs': {cnpj for cnpj in cnpjs if cnpj},
            'names': all_names,
        }

    def _extract_consolidated_row(self, row: List, column_indices: Dict[str, int]) -> Dict[str, object]:
        def get_value(column_key: str):
            idx = column_indices[column_key]
            return row[idx] if len(row) > idx else None

        alimentacao_pat = self._to_decimal(get_value('alimentacao_pat'))
        alimentacao_nao_pat = self._to_decimal(get_value('alimentacao_refeicao_nao_pat'))

        return {
            'company_name': str(get_value('company_name') or '').strip(),
            'cnpj': self._normalize_digits(get_value('cnpj')),
            'recharge_id': str(get_value('recharge_id') or '').strip(),
            'recharge_context': str(get_value('recharge_context') or '').strip(),
            'recharge_month': self._normalize_recharge_month(get_value('recharge_month')),
            'cpf': str(get_value('cpf') or '').strip(),
            'name': str(get_value('nome') or '').strip(),
            'refeicao': self._to_decimal(get_value('refeicao_pat')),
            'alimentacao': alimentacao_pat + alimentacao_nao_pat,
            'mobilidade': self._to_decimal(get_value('mobilidade')),
            'livre': self._to_decimal(get_value('livre')),
        }

    def _read_xlsx_rows(self, file_path: str) -> Tuple[List[str], List[List]]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
        return [str(h) if h is not None else "" for h in headers], rows

    def _read_csv_rows(self, file_path: str) -> Tuple[List[str], List[List]]:
        with open(file_path, 'rb') as f:
            raw = f.read()

        decoded_text = None
        decoded_encoding = None
        for encoding in ('utf-16', 'utf-16le', 'utf-16be', 'utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
            try:
                decoded_text = raw.decode(encoding)
                decoded_encoding = encoding
                break
            except UnicodeDecodeError:
                continue

        if decoded_text is None:
            decoded_text = raw.decode('latin-1', errors='replace')
            decoded_encoding = 'latin-1'

        sample = decoded_text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=';,	,')
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ';' if ';' in sample else ','

        reader = csv.reader(io.StringIO(decoded_text), delimiter=delimiter)
        rows = list(reader)

        logger.info(f"CSV decodificado usando {decoded_encoding} com delimitador '{delimiter}'")

        if not rows:
            return [], []

        headers = rows[0]
        data_rows = rows[1:]
        return headers, data_rows

    def _read_benefits_rows(self, file_path: str, file_extension: str) -> Tuple[List[str], List[List]]:
        ext = (file_extension or '').lower()
        if ext in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            return self._read_xlsx_rows(file_path)
        if ext in ['.csv', '.txt']:
            return self._read_csv_rows(file_path)
        raise ValueError(f"Extensão de arquivo não suportada: {file_extension}")

    def _build_employee_name_map(self) -> Dict[str, Employee]:
        from app.utils.parsers import normalize_name_for_payroll

        if self._employee_name_map is not None:
            return self._employee_name_map

        employees = self.db.query(Employee).all()
        name_map: Dict[str, Employee] = {}
        for employee in employees:
            normalized_name = normalize_name_for_payroll(employee.name)
            if normalized_name and normalized_name not in name_map:
                name_map[normalized_name] = employee

        self._employee_name_map = name_map
        return name_map

    @staticmethod
    def _reference_date(year: int, month: int) -> date:
        """Retorna o último dia do mês de referência."""
        return date(year, month, calendar.monthrange(year, month)[1])

    def _select_employee_by_period_context(
        self,
        candidates: List[Employee],
        year: int,
        month: int,
        company: str,
    ) -> Optional[Employee]:
        """Resolve CPF duplicado priorizando o vínculo correto no período/empresa."""
        if not candidates:
            return None
        if len(candidates) == 1:
            return candidates[0]

        candidate_ids = [employee.id for employee in candidates]

        # 1) Prioriza vínculo que aparece na folha do mesmo ano/mês/empresa.
        period_matches = self.db.query(Employee).join(
            PayrollData, PayrollData.employee_id == Employee.id
        ).join(
            PayrollPeriod, PayrollPeriod.id == PayrollData.period_id
        ).filter(
            Employee.id.in_(candidate_ids),
            PayrollPeriod.year == year,
            PayrollPeriod.month == month,
            PayrollPeriod.company == company,
        ).distinct().all()

        if len(period_matches) == 1:
            return period_matches[0]
        if len(period_matches) > 1:
            # Empate: usa o vínculo com admissão mais recente.
            return sorted(
                period_matches,
                key=lambda employee: (employee.admission_date or date.min, employee.id),
                reverse=True,
            )[0]

        # 2) Sem folha no período: limita à empresa do cadastro.
        same_company = [
            employee
            for employee in candidates
            if str(employee.company_code or '').strip() == str(company).strip()
        ]
        if len(same_company) == 1:
            return same_company[0]
        if len(same_company) > 1:
            candidates = same_company

        # 3) Prioriza vínculo ativo na data de referência do período.
        ref_date = self._reference_date(year, month)
        active_on_ref = []
        for employee in candidates:
            admission_ok = employee.admission_date is None or employee.admission_date <= ref_date
            termination_ok = employee.termination_date is None or employee.termination_date >= ref_date
            if admission_ok and termination_ok:
                active_on_ref.append(employee)

        if len(active_on_ref) == 1:
            return active_on_ref[0]
        if len(active_on_ref) > 1:
            return sorted(
                active_on_ref,
                key=lambda employee: (employee.admission_date or date.min, employee.id),
                reverse=True,
            )[0]

        # 4) Fallback final: cadastro ativo e mais recente.
        active_flag = [employee for employee in candidates if bool(employee.is_active)]
        if len(active_flag) == 1:
            return active_flag[0]
        if len(active_flag) > 1:
            candidates = active_flag

        return sorted(
            candidates,
            key=lambda employee: (employee.admission_date or date.min, employee.id),
            reverse=True,
        )[0]

    def _find_employee(
        self,
        cpf: str,
        name: Optional[str],
        row_number: int,
        year: int,
        month: int,
        company: str,
    ) -> Tuple[Optional[Employee], Optional[str]]:
        """Localiza colaborador por CPF e faz fallback para nome quando necessário."""
        from sqlalchemy import func
        from app.utils.parsers import normalize_name_for_payroll

        cpf_normalized = self.normalize_cpf(cpf)
        if cpf_normalized and len(cpf_normalized) == 11:
            cpf_candidates = self.db.query(Employee).filter(
                func.regexp_replace(Employee.cpf, r'[^0-9]', '', 'g') == cpf_normalized
            ).all()

            employee = self._select_employee_by_period_context(
                candidates=cpf_candidates,
                year=year,
                month=month,
                company=company,
            )

            if employee:
                if len(cpf_candidates) > 1:
                    self.warnings.append(
                        f"Linha {row_number}: CPF '{cpf}' com múltiplos cadastros; "
                        f"selecionado ID {employee.id} para {month:02d}/{year} empresa {company}"
                    )
                return employee, 'cpf'

        if name:
            normalized_name = normalize_name_for_payroll(name)
            if normalized_name:
                employee = self._build_employee_name_map().get(normalized_name)
                if employee:
                    warning_msg = (
                        f"Linha {row_number}: CPF '{cpf}' não casou no cadastro, "
                        f"registro associado por nome '{name}'"
                    )
                    self.warnings.append(warning_msg)
                    return employee, 'name'

        return None, None

    @staticmethod
    def _to_decimal(value) -> Decimal:
        if value is None:
            return Decimal('0')

        text = str(value).strip()
        if not text:
            return Decimal('0')

        # Remove separador de milhar e converte decimal com vírgula para ponto.
        text = text.replace('.', '').replace(',', '.') if ',' in text else text
        try:
            return Decimal(text)
        except Exception:
            return Decimal('0')
    
    @staticmethod
    def normalize_cpf(cpf: str) -> str:
        """
        Normaliza CPF removendo todos os caracteres não numéricos
        
        Args:
            cpf: CPF em qualquer formato (xxx.xxx.xxx-xx, xxx-xxx.xxx-xx, etc)
            
        Returns:
            CPF apenas com dígitos (xxxxxxxxxxx)
        """
        if not cpf:
            return ""
        # Remove tudo que não é dígito
        import re
        return re.sub(r'\D', '', str(cpf))
        
    def process_xlsx_file(
        self,
        file_path: str,
        year: int,
        month: int,
        company: str,
        period_name: Optional[str] = None,
        merge_mode: str = 'sum',
        source_label: Optional[str] = None,
        original_filename: Optional[str] = None
    ) -> Dict:
        """
        Processa arquivo XLSX de benefícios
        
        Args:
            file_path: Caminho para o arquivo XLSX
            year: Ano de referência
            month: Mês de referência
            company: Código da empresa ('0060' ou '0059')
            period_name: Nome do período (opcional, gerado automaticamente se não fornecido)
            
        Returns:
            Dicionário com resultado do processamento
        """
        start_time = datetime.now()
        self.errors = []
        self.warnings = []
        
        try:
            # Gerar nome do período se não fornecido
            if not period_name:
                month_names = [
                    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
                ]
                period_name = f"{month_names[month - 1]} {year}"
            
            # Criar ou obter período
            period = self._get_or_create_period(year, month, period_name, company)
            if not period:
                return {
                    "success": False,
                    "error": "Não foi possível criar ou encontrar o período",
                    "errors": self.errors
                }
            
            import os
            filename = original_filename or os.path.basename(file_path)
            _, file_extension = os.path.splitext(filename)

            # Carregar arquivo
            logger.info(f"Carregando arquivo de benefícios: {file_path}")
            headers, rows = self._read_benefits_rows(file_path, file_extension)
            logger.info(f"Cabeçalhos encontrados: {headers}")

            # Validação de schema estrito do novo layout consolidado
            column_indices = self._map_columns(headers)
            missing_columns = [
                key for key in self.STRICT_SCHEMA_HEADERS.keys()
                if key not in column_indices
            ]
            if missing_columns:
                error_msg = (
                    "Arquivo fora do layout consolidado iFood. "
                    f"Colunas ausentes: {', '.join(missing_columns)}"
                )
                logger.error(error_msg)
                return {
                    "success": False,
                    "error": error_msg,
                    "errors": [error_msg]
                }

            # Pré-validação e agregação por CPF
            total_rows = 0
            processed_rows = 0
            error_rows = 0
            change_set: List[Dict] = []
            matched_by_cpf = 0
            matched_by_name = 0

            period_key = f"{year}-{month:02d}"
            company_constraints = self._resolve_company_constraints(company)
            company_names_found = set()
            cnpjs_found = set()
            recharge_ids_file = set()
            validation_errors: List[str] = []

            aggregated_by_cpf = defaultdict(lambda: {
                'name': '',
                'refeicao': Decimal('0'),
                'alimentacao': Decimal('0'),
                'mobilidade': Decimal('0'),
                'livre': Decimal('0'),
                'row_numbers': [],
            })

            for row_idx, row in enumerate(rows, start=2):
                total_rows += 1
                try:
                    parsed = self._extract_consolidated_row(row, column_indices)

                    if not parsed['cpf'] and not parsed['name']:
                        continue

                    if not parsed['cpf']:
                        validation_errors.append(f"Linha {row_idx}: CPF ausente")
                        continue

                    cpf_digits = self.normalize_cpf(parsed['cpf'])
                    if len(cpf_digits) != 11:
                        validation_errors.append(f"Linha {row_idx}: CPF inválido '{parsed['cpf']}'")
                        continue

                    recharge_id = parsed['recharge_id']
                    if not recharge_id:
                        validation_errors.append(f"Linha {row_idx}: ID da recarga ausente")
                        continue

                    recharge_month = parsed['recharge_month']
                    if recharge_month != period_key:
                        validation_errors.append(
                            f"Linha {row_idx}: Mês da recarga '{recharge_month}' divergente de {period_key}"
                        )
                        continue

                    if parsed['cnpj']:
                        cnpjs_found.add(parsed['cnpj'])
                    if parsed['company_name']:
                        company_names_found.add(parsed['company_name'])

                    if company_constraints['cnpjs'] and parsed['cnpj'] not in company_constraints['cnpjs']:
                        validation_errors.append(
                            f"Linha {row_idx}: CNPJ '{parsed['cnpj']}' não pertence à empresa {company}"
                        )
                        continue

                    recharge_ids_file.add(recharge_id)

                    aggregated = aggregated_by_cpf[cpf_digits]
                    if parsed['name'] and not aggregated['name']:
                        aggregated['name'] = parsed['name']
                    aggregated['refeicao'] += parsed['refeicao']
                    aggregated['alimentacao'] += parsed['alimentacao']
                    aggregated['mobilidade'] += parsed['mobilidade']
                    aggregated['livre'] += parsed['livre']
                    aggregated['row_numbers'].append(row_idx)

                except Exception as e:
                    validation_errors.append(f"Linha {row_idx}: {str(e)}")

            if validation_errors:
                return {
                    "success": False,
                    "error": "Falha na validação do arquivo consolidado de benefícios",
                    "errors": validation_errors[:50],
                    "total_rows": total_rows,
                }

            existing_recharge_ids = self._get_existing_recharge_ids(period.id)
            duplicated_recharges = sorted(recharge_ids_file.intersection(existing_recharge_ids))
            if duplicated_recharges:
                return {
                    "success": False,
                    "error": "Upload bloqueado: existem IDs de recarga já processados para este período",
                    "errors": [f"IDs duplicados: {', '.join(duplicated_recharges[:20])}"],
                    "duplicated_recharge_ids": duplicated_recharges,
                    "total_rows": total_rows,
                }

            for cpf_digits, payload in aggregated_by_cpf.items():
                first_row = payload['row_numbers'][0] if payload['row_numbers'] else 0
                success, match_type, change_entry = self._process_benefits_row(
                    period=period,
                    cpf=cpf_digits,
                    name=payload['name'],
                    year=year,
                    month=month,
                    company=company,
                    refeicao=payload['refeicao'],
                    alimentacao=payload['alimentacao'],
                    mobilidade=payload['mobilidade'],
                    livre=payload['livre'],
                    filename=filename,
                    row_number=first_row,
                    merge_mode=merge_mode
                )

                if success and change_entry:
                    change_set.append(change_entry)

                if success:
                    processed_rows += 1
                    if match_type == 'cpf':
                        matched_by_cpf += 1
                    elif match_type == 'name':
                        matched_by_name += 1
                else:
                    error_rows += 1
            
            # Commit das alterações
            try:
                self.db.commit()
                logger.info(f"Dados de benefícios salvos com sucesso para o período {period.period_name}")
            except SQLAlchemyError as e:
                self.db.rollback()
                error_msg = f"Erro ao salvar dados: {str(e)}"
                logger.error(error_msg)
                return {
                    "success": False,
                    "error": error_msg,
                    "errors": self.errors
                }
            
            # Calcular tempo de processamento
            processing_time = (datetime.now() - start_time).total_seconds()
            
            # Criar log de processamento
            self._create_processing_log(
                period=period,
                filename=filename,
                file_size=os.path.getsize(file_path) if os.path.exists(file_path) else None,
                total_rows=total_rows,
                processed_rows=processed_rows,
                error_rows=error_rows,
                processing_time=processing_time,
                status='completed' if error_rows == 0 else 'partial',
                extra_summary={
                    "file_type": file_extension.lower().replace('.', ''),
                    "schema_version": "ifood_consolidado_v2",
                    "merge_mode": merge_mode,
                    "source_label": source_label,
                    "matched_by_cpf": matched_by_cpf,
                    "matched_by_name": matched_by_name,
                    "company": company,
                    "month": month,
                    "year": year,
                    "company_names_found": sorted(company_names_found),
                    "cnpjs_found": sorted(cnpjs_found),
                    "recharge_ids": sorted(recharge_ids_file),
                    "recharge_ids_count": len(recharge_ids_file),
                    "aggregated_rows": len(aggregated_by_cpf),
                    "rollback_version": 1,
                    "change_set": change_set,
                }
            )
            
            return {
                "success": True,
                "period_id": period.id,
                "period_name": period.period_name,
                "company": company,
                "month": month,
                "year": year,
                "source_label": source_label,
                "merge_mode": merge_mode,
                "total_rows": total_rows,
                "aggregated_rows": len(aggregated_by_cpf),
                "processed_rows": processed_rows,
                "error_rows": error_rows,
                "matched_by_cpf": matched_by_cpf,
                "matched_by_name": matched_by_name,
                "recharge_ids_count": len(recharge_ids_file),
                "recharge_ids": sorted(recharge_ids_file),
                "cnpjs_found": sorted(cnpjs_found),
                "company_names_found": sorted(company_names_found),
                "warnings": self.warnings,
                "errors": self.errors,
                "processing_time": processing_time
            }
            
        except Exception as e:
            logger.exception(f"Erro ao processar arquivo XLSX: {str(e)}")
            self.db.rollback()
            return {
                "success": False,
                "error": str(e),
                "errors": self.errors
            }
    
    def _get_or_create_period(
        self,
        year: int,
        month: int,
        period_name: str,
        company: str
    ) -> Optional[BenefitsPeriod]:
        """Busca ou cria um período de benefícios"""
        try:
            # Verificar se já existe
            period = self.db.query(BenefitsPeriod).filter(
                BenefitsPeriod.year == year,
                BenefitsPeriod.month == month,
                BenefitsPeriod.company == company
            ).first()
            
            if period:
                logger.info(f"Período existente encontrado: {period.period_name}")
                return period
            
            # Criar novo período
            period = BenefitsPeriod(
                year=year,
                month=month,
                period_name=period_name,
                company=company
            )
            self.db.add(period)
            self.db.flush()  # Para obter o ID
            
            logger.info(f"Novo período criado: {period.period_name}")
            return period
            
        except Exception as e:
            error_msg = f"Erro ao criar período: {str(e)}"
            logger.error(error_msg)
            self.errors.append(error_msg)
            return None
    
    def _process_benefits_row(
        self,
        period: BenefitsPeriod,
        cpf: str,
        name: str,
        year: int,
        month: int,
        company: str,
        refeicao,
        alimentacao,
        mobilidade,
        livre,
        filename: str,
        row_number: int,
        merge_mode: str = 'sum'
    ) -> Tuple[bool, Optional[str], Optional[Dict]]:
        """Processa uma linha de benefícios"""
        try:
            employee, match_type = self._find_employee(
                cpf=cpf,
                name=name,
                row_number=row_number,
                year=year,
                month=month,
                company=company,
            )

            if not employee:
                warning_msg = (
                    f"Linha {row_number}: colaborador não encontrado "
                    f"(CPF='{cpf}' Nome='{name}')"
                )
                logger.warning(warning_msg)
                self.warnings.append(warning_msg)
                return False, None, None
            
            # Verificar se já existe registro para este período e colaborador
            existing = self.db.query(BenefitsData).filter(
                BenefitsData.period_id == period.id,
                BenefitsData.employee_id == employee.id
            ).first()

            refeicao_value = self._to_decimal(refeicao)
            alimentacao_value = self._to_decimal(alimentacao)
            mobilidade_value = self._to_decimal(mobilidade)
            livre_value = self._to_decimal(livre)
            
            if existing:
                before_snapshot = {
                    'refeicao': str(existing.refeicao or Decimal('0')),
                    'alimentacao': str(existing.alimentacao or Decimal('0')),
                    'mobilidade': str(existing.mobilidade or Decimal('0')),
                    'livre': str(existing.livre or Decimal('0')),
                    'upload_filename': existing.upload_filename,
                    'processed_by': existing.processed_by,
                }
                # Atualizar registro existente
                if merge_mode == 'replace':
                    existing.refeicao = refeicao_value
                    existing.alimentacao = alimentacao_value
                    existing.mobilidade = mobilidade_value
                    existing.livre = livre_value
                else:
                    existing.refeicao = (existing.refeicao or Decimal('0')) + refeicao_value
                    existing.alimentacao = (existing.alimentacao or Decimal('0')) + alimentacao_value
                    existing.mobilidade = (existing.mobilidade or Decimal('0')) + mobilidade_value
                    existing.livre = (existing.livre or Decimal('0')) + livre_value
                existing.upload_filename = filename
                existing.processed_by = self.user_id

                after_snapshot = {
                    'refeicao': str(existing.refeicao or Decimal('0')),
                    'alimentacao': str(existing.alimentacao or Decimal('0')),
                    'mobilidade': str(existing.mobilidade or Decimal('0')),
                    'livre': str(existing.livre or Decimal('0')),
                    'upload_filename': existing.upload_filename,
                    'processed_by': existing.processed_by,
                }

                change_entry = {
                    'action': 'updated',
                    'employee_id': employee.id,
                    'benefits_data_id': existing.id,
                    'before': before_snapshot,
                    'after': after_snapshot,
                }
            else:
                # Criar novo registro
                benefits_data = BenefitsData(
                    period_id=period.id,
                    employee_id=employee.id,
                    cpf=cpf,
                    refeicao=refeicao_value,
                    alimentacao=alimentacao_value,
                    mobilidade=mobilidade_value,
                    livre=livre_value,
                    upload_filename=filename,
                    processed_by=self.user_id
                )
                self.db.add(benefits_data)

                change_entry = {
                    'action': 'created',
                    'employee_id': employee.id,
                    'before': None,
                    'after': {
                        'refeicao': str(refeicao_value),
                        'alimentacao': str(alimentacao_value),
                        'mobilidade': str(mobilidade_value),
                        'livre': str(livre_value),
                        'upload_filename': filename,
                        'processed_by': self.user_id,
                    },
                }
            
            return True, match_type, change_entry
            
        except Exception as e:
            error_msg = f"Linha {row_number}: {str(e)}"
            logger.error(error_msg)
            self.errors.append(error_msg)
            return False, None, None

    def _get_existing_recharge_ids(self, period_id: int) -> set:
        logs = self.db.query(BenefitsProcessingLog).filter(
            BenefitsProcessingLog.period_id == period_id
        ).all()

        recharge_ids = set()
        for log in logs:
            summary = log.processing_summary or {}
            log_recharge_ids = summary.get('recharge_ids') or []
            if isinstance(log_recharge_ids, list):
                recharge_ids.update(str(item).strip() for item in log_recharge_ids if str(item).strip())

        return recharge_ids
    
    def _create_processing_log(
        self,
        period: BenefitsPeriod,
        filename: str,
        file_size: Optional[int],
        total_rows: int,
        processed_rows: int,
        error_rows: int,
        processing_time: float,
        status: str,
        extra_summary: Optional[Dict] = None
    ):
        """Cria log de processamento"""
        try:
            summary = {
                "warnings": self.warnings,
                "errors": self.errors
            }
            if extra_summary:
                summary.update(extra_summary)

            log = BenefitsProcessingLog(
                period_id=period.id,
                filename=filename,
                file_size=file_size,
                total_rows=total_rows,
                processed_rows=processed_rows,
                error_rows=error_rows,
                status=status,
                processing_summary=summary,
                processed_by=self.user_id,
                processing_time=Decimal(str(round(processing_time, 2)))
            )
            self.db.add(log)
            self.db.commit()
            logger.info(f"Log de processamento criado para {filename}")
        except Exception as e:
            logger.error(f"Erro ao criar log de processamento: {str(e)}")
