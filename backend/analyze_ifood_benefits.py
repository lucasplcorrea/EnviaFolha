"""
Script para analisar estrutura do arquivo de benefícios iFood
"""
import openpyxl
from pathlib import Path

xlsx_path = Path(__file__).parent.parent / "Analiticos" / "Ifood" / "Layout ifood Refeição 012026.xlsx"

# Carregar o arquivo
wb = openpyxl.load_workbook(xlsx_path)
print(f"Abas disponíveis: {wb.sheetnames}")

# Pegar a primeira aba
ws = wb.active
print(f"\nAba ativa: {ws.title}")

# Mostrar cabeçalhos (primeira linha)
headers = []
for cell in ws[1]:
    headers.append(cell.value)
print(f"\nCabeçalhos ({len(headers)} colunas):")
for i, header in enumerate(headers, 1):
    print(f"  {i}. {header}")

# Mostrar primeiras 5 linhas de dados
print(f"\nPrimeiras 5 linhas de dados:")
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 2):
    print(f"\nLinha {i}:")
    for header, value in zip(headers, row):
        if value is not None:
            print(f"  {header}: {value}")

# Contar total de linhas
total_rows = ws.max_row
print(f"\nTotal de linhas (incluindo cabeçalho): {total_rows}")
print(f"Total de linhas de dados: {total_rows - 1}")
