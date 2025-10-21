"""
Script para criar arquivo Excel modelo com formatação correta
"""
import sys
import os

backend_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, backend_dir)

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    AVAILABLE = True
except:
    AVAILABLE = False
    print("❌ pandas ou openpyxl não disponíveis")
    exit(1)

def create_template():
    print("📝 Criando arquivo modelo de importação...\n")
    
    # Dados de exemplo
    data = {
        'unique_id': ['0060XXXXX', '0059XXXXX', '0001234', '0000001'],  # Com zeros à esquerda
        'full_name': ['João da Silva', 'Maria Santos', 'Pedro Oliveira', 'Ana Costa'],
        'cpf': ['12345678901', '98765432100', '11122233344', '55566677788'],  # 11 dígitos
        'phone_number': ['47999999999', '11988888888', '47988887777', '11977776666'],  # Com DDD
        'email': ['joao@example.com', 'maria@example.com', '', ''],  # Opcional
        'department': ['TI', 'RH', 'Vendas', ''],  # Opcional
        'position': ['Analista', 'Coordenadora', 'Vendedor', ''],  # Opcional
    }
    
    df = pd.DataFrame(data)
    
    # Salvar Excel
    output_file = 'Modelo_Importacao_Colaboradores.xlsx'
    
    # Usar ExcelWriter para mais controle
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Colaboradores', index=False)
        
        # Acessar worksheet para formatação
        worksheet = writer.sheets['Colaboradores']
        
        # Formatar cabeçalho
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # FORÇAR FORMATO TEXTO para colunas críticas
        for col_idx, col_name in enumerate(['unique_id', 'cpf', 'phone_number'], start=1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            for row in range(2, len(df) + 2):  # Pular cabeçalho
                cell = worksheet[f'{col_letter}{row}']
                cell.number_format = '@'  # '@' = formato texto no Excel
        
        # Ajustar largura das colunas
        worksheet.column_dimensions['A'].width = 15  # unique_id
        worksheet.column_dimensions['B'].width = 30  # full_name
        worksheet.column_dimensions['C'].width = 15  # cpf
        worksheet.column_dimensions['D'].width = 15  # phone_number
        worksheet.column_dimensions['E'].width = 30  # email
        worksheet.column_dimensions['F'].width = 15  # department
        worksheet.column_dimensions['G'].width = 20  # position
        
    print(f"✅ Arquivo criado: {output_file}")
    print(f"\n📋 Colunas obrigatórias (com formatação TEXTO):")
    print("   - unique_id: Código único do colaborador (ex: 0060XXXXX)")
    print("   - full_name: Nome completo")
    print("   - cpf: CPF com 11 dígitos (sem formatação)")
    print("   - phone_number: Telefone com DDD (ex: 47999999999)")
    print(f"\n📋 Colunas opcionais:")
    print("   - email, department, position")
    print(f"\n⚠️ IMPORTANTE:")
    print("   As colunas unique_id, cpf e phone_number já estão")
    print("   formatadas como TEXTO para preservar zeros à esquerda!")
    
    return True

if __name__ == "__main__":
    if AVAILABLE:
        create_template()
    else:
        print("Instale pandas e openpyxl: pip install pandas openpyxl")
