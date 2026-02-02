"""
Análise da planilha de Cartão Ponto
"""
import openpyxl
import os

# Caminho do arquivo
xlsx_path = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "Analiticos", "Cartão Ponto", "ExtratoTotais_21.02.25-20.03.25_AB.xlsx"
)

print("=" * 80)
print("📊 ANÁLISE DE PLANILHA DE CARTÃO PONTO")
print("=" * 80)
print(f"📁 Arquivo: {os.path.basename(xlsx_path)}")
print()

# Carregar workbook
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

print(f"📑 Abas disponíveis: {wb.sheetnames}")
print()

# Analisar primeira aba
sheet = wb.active
print(f"🔍 Analisando aba: {sheet.title}")
print("=" * 80)

# Encontrar linha de headers (procurar por "Nº Folha" ou "Nome")
header_row = 1
for row_idx in range(1, min(10, sheet.max_row + 1)):
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row_idx, col_idx).value
        if cell_value and isinstance(cell_value, str):
            if 'Nº Folha' in cell_value or 'Nome' in cell_value and 'Ex' in str(sheet.cell(row_idx, col_idx + 2).value or ''):
                header_row = row_idx
                print(f"✅ Headers encontrados na linha {header_row}")
                break
    if header_row > 1:
        break

print()
print(f"📄 PRIMEIRAS 10 LINHAS (RAW):")
print("-" * 80)
for row_idx in range(1, min(11, sheet.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(12, sheet.max_column + 1)):
        val = sheet.cell(row_idx, col_idx).value
        row_data.append(str(val)[:15] if val else "")
    print(f"Linha {row_idx:2d}: {' | '.join(row_data)}")

print()
print(f"📋 Usando linha {header_row} como headers")
print()

# Ler headers da linha correta
headers = []
for col in range(1, sheet.max_column + 1):
    cell_value = sheet.cell(header_row, col).value
    headers.append(cell_value)

print(f"\n📋 COLUNAS ENCONTRADAS ({len(headers)}):")
print("-" * 80)
for i, header in enumerate(headers, 1):
    print(f"{i:2d}. {header}")

# Contar linhas de dados
total_rows = sheet.max_row
data_rows = total_rows - header_row  # Excluindo header

print()
print(f"📊 ESTATÍSTICAS:")
print("-" * 80)
print(f"• Total de linhas (incluindo header): {total_rows}")
print(f"• Linhas de dados: {data_rows}")
print(f"• Total de colunas: {len(headers)}")

# Mostrar primeiras 5 linhas de dados
print()
print("📝 PRIMEIRAS 5 LINHAS DE DADOS:")
print("-" * 80)

# Criar formato de tabela
col_widths = [max(15, len(str(h)) if h else 10) for h in headers]

# Header
header_line = " | ".join(f"{str(h)[:15]:15s}" if h else " "*15 for h in headers)
print(header_line)
print("-" * len(header_line))

# Dados
for row_idx in range(header_row + 1, min(header_row + 6, total_rows + 1)):  # Primeiras 5 linhas
    row_data = []
    for col_idx in range(1, len(headers) + 1):
        cell_value = sheet.cell(row_idx, col_idx).value
        value_str = str(cell_value) if cell_value is not None else ""
        row_data.append(f"{value_str[:15]:15s}")
    
    print(" | ".join(row_data))

# Análise específica das colunas relevantes
print()
print("🎯 ANÁLISE DE COLUNAS RELEVANTES:")
print("-" * 80)

relevant_cols = {
    'Nº Folha': None,  # Matrícula
    'Nome': None,
    'Normais': None,
    'Ex50%': None,
    'Ex100%': None,
    'EN50%': None,
    'EN100%': None,
    'Not.': None,
    'Faltas': None,
    'DSR.Deb': None,
    'Abono2': None
}

# Encontrar índices das colunas (match exato)
for i, header in enumerate(headers, 1):
    header_str = str(header).strip() if header else ""
    if header_str in relevant_cols:
        relevant_cols[header_str] = i

print("Colunas mapeadas:")
for col_name, col_idx in relevant_cols.items():
    status = "✅ ENCONTRADA" if col_idx else "❌ NÃO ENCONTRADA"
    col_letter = f"Coluna {col_idx}" if col_idx else "N/A"
    print(f"  • {col_name:20s}: {status:20s} ({col_letter})")

# Análise de valores das horas extras
print()
print("📊 ANÁLISE DE VALORES (Primeiras 10 linhas):")
print("-" * 80)

if relevant_cols['Nº Folha'] and relevant_cols['Nome']:
    print(f"{'Nº Folha':<12} {'Nome':<30} {'Ex50%':<15} {'Ex100%':<15} {'EN50%':<15} {'EN100%':<15} {'Not.':<15}")
    print("-" * 100)
    
    for row_idx in range(header_row + 1, min(header_row + 11, total_rows + 1)):
        matricula = sheet.cell(row_idx, relevant_cols['Nº Folha']).value if relevant_cols['Nº Folha'] else ""
        nome = sheet.cell(row_idx, relevant_cols['Nome']).value if relevant_cols['Nome'] else ""
        ex50 = sheet.cell(row_idx, relevant_cols['Ex50%']).value if relevant_cols['Ex50%'] else 0
        ex100 = sheet.cell(row_idx, relevant_cols['Ex100%']).value if relevant_cols['Ex100%'] else 0
        en50 = sheet.cell(row_idx, relevant_cols['EN50%']).value if relevant_cols['EN50%'] else 0
        en100 = sheet.cell(row_idx, relevant_cols['EN100%']).value if relevant_cols['EN100%'] else 0
        not_ = sheet.cell(row_idx, relevant_cols['Not.']).value if relevant_cols['Not.'] else 0
        
        # Converter para string, limitando tamanho
        nome_str = str(nome)[:28] if nome else ""
        
        print(f"{str(matricula):<12} {nome_str:<30} {str(ex50):<15} {str(ex100):<15} {str(en50):<15} {str(en100):<15} {str(not_):<15}")

# Estatísticas de horas
print()
print("📈 ESTATÍSTICAS DE HORAS:")
print("-" * 80)

if relevant_cols['Ex50%']:
    total_ex50 = 0
    count_ex50 = 0
    for row_idx in range(header_row + 1, total_rows + 1):
        val = sheet.cell(row_idx, relevant_cols['Ex50%']).value
        if val and isinstance(val, (int, float)) and val > 0:
            total_ex50 += val
            count_ex50 += 1
    
    print(f"• Ex50% (Extras 50%):")
    print(f"  - Total de horas: {total_ex50}")
    print(f"  - Colaboradores com extras: {count_ex50}")
    if count_ex50 > 0:
        print(f"  - Média por colaborador: {total_ex50/count_ex50:.2f}h")

if relevant_cols['Ex100%']:
    total_ex100 = 0
    count_ex100 = 0
    for row_idx in range(header_row + 1, total_rows + 1):
        val = sheet.cell(row_idx, relevant_cols['Ex100%']).value
        if val and isinstance(val, (int, float)) and val > 0:
            total_ex100 += val
            count_ex100 += 1
    
    print(f"\n• Ex100% (Extras 100%):")
    print(f"  - Total de horas: {total_ex100}")
    print(f"  - Colaboradores com extras: {count_ex100}")
    if count_ex100 > 0:
        print(f"  - Média por colaborador: {total_ex100/count_ex100:.2f}h")

if relevant_cols['Not.']:
    total_not = 0
    count_not = 0
    for row_idx in range(header_row + 1, total_rows + 1):
        val = sheet.cell(row_idx, relevant_cols['Not.']).value
        if val and isinstance(val, (int, float)) and val > 0:
            total_not += val
            count_not += 1
    
    print(f"\n• Not. (Noturnas):")
    print(f"  - Total de horas: {total_not}")
    print(f"  - Colaboradores com noturnas: {count_not}")
    if count_not > 0:
        print(f"  - Média por colaborador: {total_not/count_not:.2f}h")

print()
print("=" * 80)
print("✅ Análise concluída!")
print("=" * 80)

wb.close()
