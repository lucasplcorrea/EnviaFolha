"""
Gera o modelo de importação de colaboradores em xlsx com instruções e validações.
"""
import sys
sys.path.insert(0, '.')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl não instalado. Instale com: pip install openpyxl")
    sys.exit(1)

wb = openpyxl.Workbook()

# ─── Aba 1: Dados ───────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Colaboradores"

# Estilos
header_obrig = PatternFill("solid", fgColor="1F4E79")  # azul escuro - obrigatório
header_opcio = PatternFill("solid", fgColor="2E75B6")  # azul médio - opcional
font_header = Font(color="FFFFFF", bold=True, size=10)
font_example = Font(color="595959", italic=True, size=9)
font_note = Font(color="B00020", bold=True, size=9)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLUMNS = [
    # (header, example, required, width, note)
    ("nome",          "João da Silva",       True,  30, ""),
    ("cpf",           "04775016997",         True,  16, "Somente números (11 dígitos)"),
    ("matricula",     "571",                 True,  12, "Número da matrícula do colaborador"),
    ("company_code",  "0059",               True,  14, "0059=Infraestrutura / 0060=Empreendimentos"),
    ("data_admissao", "01-06-2025",          True,  16, "Formato: DD-MM-AAAA"),
    ("cargo",         "Servente de Obras",   False, 28, ""),
    ("departamento",  "Obras",               False, 22, ""),
    ("sexo",          "M",                   False, 8,  "M ou F"),
    ("data_nascimento","07-10-1982",         False, 16, "Formato: DD-MM-AAAA"),
    ("estado_civil",  "Solteiro",            False, 16, ""),
    ("tipo_contrato", "CLT",                 False, 14, "CLT, PJ, Estagiário, etc."),
    ("telefone",      "5546988237067",       False, 18, "Com código do país (55) e DDD"),
    ("email",         "joao@empresa.com.br", False, 28, ""),
    ("situacao",      "Ativo",               False, 16, "Ativo, Férias, Afastado, Desligado"),
    ("data_demissao", "",                    False, 16, "Preencher apenas se Demitido. Formato: DD-MM-AAAA"),
]

# Linha 1: legenda de cores
ws.row_dimensions[1].height = 18
ws.merge_cells("A1:D1")
c = ws["A1"]; c.value = "🔵 Obrigatório"; c.font = Font(color="FFFFFF", bold=True, size=9)
c.fill = header_obrig; c.alignment = align_center
ws.merge_cells("E1:O1")
c = ws["E1"]; c.value = "⚪ Opcional"
c.font = Font(color="FFFFFF", bold=True, size=9)
c.fill = header_opcio; c.alignment = align_center

# Linha 2: cabeçalho das colunas
ws.row_dimensions[2].height = 30
for col_idx, (col_name, _, required, width, _) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=2, column=col_idx, value=col_name)
    cell.font = font_header
    cell.fill = header_obrig if required else header_opcio
    cell.alignment = align_center
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Linha 3: exemplos
ws.row_dimensions[3].height = 18
for col_idx, (_, example, _, _, _) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=3, column=col_idx, value=example)
    cell.font = font_example
    cell.fill = PatternFill("solid", fgColor="EBF3FB")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Linhas 4-103: área de preenchimento (zebrada)
fill_even = PatternFill("solid", fgColor="FFFFFF")
fill_odd  = PatternFill("solid", fgColor="F5FAFF")
for row in range(4, 104):
    ws.row_dimensions[row].height = 16
    fill = fill_odd if row % 2 == 0 else fill_even
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Linha de notas (logo abaixo do cabeçalho, linha 4 especial)
ws.freeze_panes = "A4"

# ─── Aba 2: Instruções ───────────────────────────────────────────────────────
wi = wb.create_sheet("Instruções")
wi.column_dimensions["A"].width = 25
wi.column_dimensions["B"].width = 60
wi.column_dimensions["C"].width = 30

wi["A1"] = "GUIA DE PREENCHIMENTO — Importação de Colaboradores"
wi["A1"].font = Font(bold=True, size=14, color="1F4E79")
wi.merge_cells("A1:C1")

rows_inst = [
    ("", "", ""),
    ("CAMPO", "DESCRIÇÃO", "OBSERVAÇÕES"),
    ("nome", "Nome completo do colaborador", "Obrigatório"),
    ("cpf", "CPF somente números, sem pontos ou traço", "Obrigatório — 11 dígitos"),
    ("matricula", "Número da matrícula do colaborador", "Obrigatório — número puro, ex: 571"),
    ("company_code", "Código da empresa no sistema de folha", "Obrigatório — 0059 ou 0060"),
    ("data_admissao", "Data de admissão no emprego", "Obrigatório — AAAA-MM-DD ou DD/MM/AAAA"),
    ("cargo", "Cargo ou função exercida", "Opcional"),
    ("departamento", "Departamento do colaborador", "Opcional"),
    ("setor", "Setor ou área de atuação", "Opcional"),
    ("sexo", "Gênero", "Opcional — M ou F"),
    ("data_nascimento", "Data de nascimento", "Opcional — AAAA-MM-DD ou DD/MM/AAAA"),
    ("estado_civil", "Estado civil", "Opcional — Solteiro, Casado, etc."),
    ("tipo_contrato", "Tipo de contrato de trabalho", "Opcional — CLT, PJ, Estagiário..."),
    ("telefone", "Telefone celular com código do país", "Opcional — ex: 5546988237067"),
    ("email", "Endereço de e-mail do colaborador", "Opcional"),
    ("situacao", "Situação atual do colaborador", "Opcional — Ativo, Férias, Afastado, Desligado"),
    ("", "", ""),
    ("⚠️  ATENÇÃO", "O campo absolute_id é gerado AUTOMATICAMENTE pelo sistema.", "Não inclua essa coluna."),
    ("⚠️  ATENÇÃO", "A linha 3 da aba Colaboradores contém EXEMPLOS. Apague-a antes de importar, ou deixe-a, o sistema ignorará valores repetidos de example.", ""),
    ("⚠️  ATENÇÃO", "Se o mesmo colaborador for importado mais de uma vez com cpf + matricula + empresa iguais, o registro será ATUALIZADO, não duplicado.", ""),
]

header_fill = PatternFill("solid", fgColor="1F4E79")
note_fill   = PatternFill("solid", fgColor="FFF3CD")
for r_idx, (a, b, c_val) in enumerate(rows_inst, start=2):
    ca = wi.cell(row=r_idx, column=1, value=a)
    cb = wi.cell(row=r_idx, column=2, value=b)
    cc = wi.cell(row=r_idx, column=3, value=c_val)
    if a == "CAMPO":
        for cell in (ca, cb, cc):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = align_center
    elif a.startswith("⚠️"):
        for cell in (ca, cb, cc):
            cell.fill = note_fill
            cell.font = Font(bold=True, color="B00020", size=9)
    for cell in (ca, cb, cc):
        cell.border = border
        if not cell.alignment or cell.alignment.horizontal is None:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    if r_idx > 3:
        wi.row_dimensions[r_idx].height = 20

output = "modelo_importacao_colaboradores.xlsx"
wb.save(output)
print(f"✅ Modelo gerado: {output}")
